import asyncio
import os
import sys
from pathlib import Path
from typing import Dict, Any
import math

import pandas as pd

# Ensure project root is on sys.path so this script works when executed directly
PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from nep_postgres_client import DB_CONFIG  # use NEP DB config
import asyncpg
from openpyxl import load_workbook


EXCEL_PATH = os.getenv(
    "PBC_GAB_2026_EXCEL_PATH",
    "/home/joebert/open-data-visualization/database/[PBC] 2026 Budget Data Analysis.xlsx",
)

TABLE_NAME = "pbc_gab_2026_rows"
HEADINGS_TABLE = "pbc_gab_2026_headings"
HEADINGS_DETAIL_TABLE = "pbc_gab_2026_headings_detail"


async def ensure_table(conn: asyncpg.Connection) -> None:
    await conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
            id SERIAL PRIMARY KEY,
            sheet_name TEXT NOT NULL,
            row_index INTEGER NOT NULL,
            data JSONB NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_sheet ON {TABLE_NAME}(sheet_name);
        CREATE INDEX IF NOT EXISTS idx_{TABLE_NAME}_row ON {TABLE_NAME}(row_index);

        CREATE TABLE IF NOT EXISTS {HEADINGS_TABLE} (
            id SERIAL PRIMARY KEY,
            sheet_name TEXT NOT NULL,
            label TEXT NOT NULL,
            data JSONB NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_{HEADINGS_TABLE}_sheet ON {HEADINGS_TABLE}(sheet_name);

        CREATE TABLE IF NOT EXISTS {HEADINGS_DETAIL_TABLE} (
            id SERIAL PRIMARY KEY,
            sheet_name TEXT NOT NULL,
            label TEXT NOT NULL,
            original NUMERIC,
            hgab NUMERIC,
            delta NUMERIC
        );
        CREATE INDEX IF NOT EXISTS idx_{HEADINGS_DETAIL_TABLE}_sheet ON {HEADINGS_DETAIL_TABLE}(sheet_name);
        """
    )


def _clean_value(v: Any) -> Any:
    try:
        if v is None:
            return None
        # Handle pandas/numpy NaN
        if isinstance(v, float) and math.isnan(v):
            return None
        # Strings stay as-is
        return v
    except Exception:
        return None


def dataframe_to_records(df: pd.DataFrame) -> Dict[int, Dict[str, Any]]:
    # Replace pandas NaN/NA with None
    df = df.where(pd.notnull(df), None)
    raw = df.to_dict(orient="index")
    cleaned: Dict[int, Dict[str, Any]] = {}
    for i, row in raw.items():
        cleaned[int(i)] = {str(k): _clean_value(v) for k, v in row.items()}
    return cleaned


async def load_sheet(conn: asyncpg.Connection, sheet_name: str, df: pd.DataFrame) -> int:
    records = dataframe_to_records(df)
    if not records:
        return 0
    await conn.execute("DELETE FROM " + TABLE_NAME + " WHERE sheet_name=$1", sheet_name)
    import json as _json
    rows = [(sheet_name, idx, _json.dumps(data, allow_nan=False)) for idx, data in records.items()]
    await conn.executemany(
        "INSERT INTO " + TABLE_NAME + " (sheet_name, row_index, data) VALUES ($1, $2, $3::jsonb)",
        rows,
    )
    return len(rows)


def extract_headings_from_first_sheet(excel_path: str) -> tuple[str, list[dict]]:
    wb = load_workbook(excel_path, data_only=True)
    first_sheet = wb.sheetnames[0]
    ws = wb[first_sheet]

    # Determine header row: first row with at least 2 non-empty cells
    header = []
    header_row_idx = None
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 10)):
        values = [cell.value for cell in row]
        non_empty = [v for v in values if v is not None and str(v).strip() != ""]
        if len(non_empty) >= 2:
            header = [str(v).strip() if v is not None else "" for v in values]
            header_row_idx = row[0].row
            break
    if not header:
        header = []
        header_row_idx = 1

    def is_bold_underline(cell) -> bool:
        f = cell.font
        if not f:
            return False
        is_bold = bool(f.bold)
        is_underline = bool(f.underline)
        return is_bold and is_underline

    results: list[dict] = []
    # Iterate rows after header
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        first_cell = row[0]
        if not first_cell.value:
            continue
        if not is_bold_underline(first_cell):
            continue
        label = str(first_cell.value).strip()
        data: dict[str, float] = {}
        for idx, cell in enumerate(row[1:], start=1):
            raw = cell.value
            if raw is None or str(raw).strip() == "":
                continue
            try:
                n = float(raw)
            except Exception:
                # try to parse numbers with commas
                try:
                    n = float(str(raw).replace(",", "").strip())
                except Exception:
                    continue
            key = header[idx] if idx < len(header) and header[idx] else f"col_{idx+1}"
            data[key] = data.get(key, 0.0) + n
        if data:
            results.append({"label": label, "data": data})

    return first_sheet, results


async def store_headings(conn: asyncpg.Connection, sheet_name: str, items: list[dict]) -> int:
    await conn.execute("DELETE FROM " + HEADINGS_TABLE + " WHERE sheet_name=$1", sheet_name)
    if not items:
        return 0
    import json as _json
    rows = [(sheet_name, it["label"], _json.dumps(it["data"])) for it in items]
    await conn.executemany(
        "INSERT INTO " + HEADINGS_TABLE + " (sheet_name, label, data) VALUES ($1, $2, $3::jsonb)",
        rows,
    )
    return len(rows)


def parse_value_number(s: Any) -> float | None:
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s)
    try:
        # Handle formats like 4.402.808.041.000
        cleaned = str(s)
        cleaned = cleaned.replace("\u00a0", " ").strip()
        # remove all non-digit except minus
        cleaned = "".join(ch for ch in cleaned if ch.isdigit() or ch == '-')
        if cleaned in ("", "-"):
            return None
        return float(cleaned)
    except Exception:
        return None


def extract_headings_detail_first_sheet(excel_path: str) -> tuple[str, list[dict]]:
    wb = load_workbook(excel_path, data_only=True)
    first_sheet = wb.sheetnames[0]
    ws = wb[first_sheet]

    # Find header row and target columns
    header_row_idx = None
    header_map: dict[int, str] = {}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20)):
        values = [cell.value for cell in row]
        text_values = [str(v).strip() if v is not None else "" for v in values]
        if any("GAB Original" in tv for tv in text_values) and any("3rd Reading" in tv for tv in text_values):
            header_row_idx = row[0].row
            for idx, tv in enumerate(text_values):
                if tv:
                    header_map[idx] = tv
            break

    if header_row_idx is None:
        # fallback: assume columns A, B, C, D are label/original/hgab/delta
        header_row_idx = 1
        header_map = {0: 'Label', 1: 'GAB Original', 2: '3rd Reading HGAB', 3: 'Change from GAB to 3rd Reading Bill'}

    # Determine column indices
    def find_col(name: str) -> int | None:
        name_lower = name.lower()
        for idx, tv in header_map.items():
            t = tv.lower()
            if name_lower in t:
                return idx
        return None

    col_label = 0
    col_original = find_col('GAB Original') or 1
    col_hgab = find_col('3rd Reading') or 2
    col_delta = find_col('Change from GAB') or 3

    results: list[dict] = []
    for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=ws.max_row):
        label_cell = row[col_label]
        if label_cell.value is None or str(label_cell.value).strip() == "":
            continue
        label = str(label_cell.value).strip()
        v_original = parse_value_number(row[col_original].value) if col_original < len(row) else None
        v_hgab = parse_value_number(row[col_hgab].value) if col_hgab < len(row) else None
        v_delta = parse_value_number(row[col_delta].value) if col_delta < len(row) else None
        # accept rows that have at least one numeric value
        if any(v is not None for v in (v_original, v_hgab, v_delta)):
            results.append({
                'label': label,
                'original': v_original,
                'hgab': v_hgab,
                'delta': v_delta,
            })

    return first_sheet, results


async def store_headings_detail(conn: asyncpg.Connection, sheet_name: str, items: list[dict]) -> int:
    await conn.execute("DELETE FROM " + HEADINGS_DETAIL_TABLE + " WHERE sheet_name=$1", sheet_name)
    if not items:
        return 0
    rows = [(sheet_name, it['label'], it.get('original'), it.get('hgab'), it.get('delta')) for it in items]
    await conn.executemany(
        "INSERT INTO " + HEADINGS_DETAIL_TABLE + " (sheet_name, label, original, hgab, delta) VALUES ($1, $2, $3, $4, $5)",
        rows,
    )
    return len(rows)


async def main() -> None:
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel file not found: {EXCEL_PATH}")

    xls = pd.ExcelFile(EXCEL_PATH)
    conn = await asyncpg.connect(**DB_CONFIG)
    try:
        await ensure_table(conn)
        total_inserted = 0
        for sheet in xls.sheet_names:
            df = xls.parse(sheet)
            inserted = await load_sheet(conn, sheet, df)
            total_inserted += inserted
            print(f"Loaded {inserted} rows from sheet: {sheet}")

        # Extract heading departments from first sheet (bold+underline rows)
        first_sheet, headings = extract_headings_from_first_sheet(EXCEL_PATH)
        hcount = await store_headings(conn, first_sheet, headings)
        print(f"Stored {hcount} heading rows from first sheet: {first_sheet}")

        # Extract detailed columns for first sheet
        d_first_sheet, headings_detail = extract_headings_detail_first_sheet(EXCEL_PATH)
        dhcount = await store_headings_detail(conn, d_first_sheet, headings_detail)
        print(f"Stored {dhcount} heading detail rows from first sheet: {d_first_sheet}")
        print(f"Done. Inserted total rows: {total_inserted}")
    finally:
        await conn.close()


if __name__ == "__main__":
    asyncio.run(main())


