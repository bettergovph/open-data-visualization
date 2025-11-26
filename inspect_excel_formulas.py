#!/usr/bin/env python3
"""
Inspect Excel formulas and structure in General Summary file
"""

import openpyxl
from pathlib import Path

file_path = Path("database/win/General Summary of Committee Report No. 18 on House Bill No. 4058 (FY 2026 GAB).xlsx")

print("🔍 Inspecting General Summary Excel File\n")
print("="*80)

# Load workbook with formulas
wb = openpyxl.load_workbook(file_path, data_only=False)
ws = wb['By Agency']

print(f"Sheet: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print("\n" + "="*80 + "\n")

# Look at rows around the data (skip headers)
print("Inspecting rows 35-45 (around row 43):\n")

for row_num in range(35, 46):
    row = ws[row_num]
    
    # Get first column value
    first_col_value = row[0].value
    
    # Check if it has formulas
    has_formulas = any(cell.data_type == 'f' for cell in row)
    
    # Get a sample formula if exists
    sample_formula = None
    for cell in row:
        if cell.data_type == 'f':
            sample_formula = f"{cell.coordinate}: {cell.value}"
            break
    
    print(f"Row {row_num}: [{first_col_value}]")
    if has_formulas:
        print(f"  ✓ Has formulas - Example: {sample_formula}")
    print()

print("="*80)
print("\nLooking at column A (department names) rows 1-50:\n")

for row_num in range(1, 51):
    cell_a = ws[f'A{row_num}']
    if cell_a.value:
        cell_type = "FORMULA" if cell_a.data_type == 'f' else "VALUE"
        print(f"Row {row_num}: [{cell_type}] {cell_a.value}")

print("\n" + "="*80)
print("\nChecking row 43 specifically:\n")

row_43 = ws[43]
for idx, cell in enumerate(row_43[:10], start=1):
    col_letter = openpyxl.utils.get_column_letter(idx)
    cell_ref = f"{col_letter}43"
    
    if cell.data_type == 'f':
        print(f"{cell_ref}: [FORMULA] {cell.value}")
    elif cell.value is not None:
        print(f"{cell_ref}: [VALUE] {cell.value}")

wb.close()
