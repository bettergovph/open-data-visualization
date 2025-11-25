#!/usr/bin/env python3
"""
Budget Amendment Excel Parser
Parses 6 Excel files from database/win/ and generates unified JSON for FY 2026 amendments.

Usage:
    python3 scripts/parse_budget_amendments.py

Output:
    static/data/budget_amendments_2026.json
"""

import pandas as pd
import json
import openpyxl
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any
import re
import asyncio
import asyncpg
from difflib import SequenceMatcher

class BudgetAmendmentParser:
    def __init__(self, base_dir: str = "database/win"):
        self.base_dir = Path(base_dir)
        self.departments = []
        self.programs = []
        self.projects = []
        self.line_items = []  # Detailed line-by-line amendments
        self.worksheets = []  # Worksheet entries (one per sheet)
        self.dept_counter = 0
        self.prog_counter = 0
        self.proj_counter = 0
        self.line_item_counter = 0
    
    def col_num_to_letter(self, n):
        """Convert column number to Excel letter (1->A, 27->AA)"""
        result = ""
        while n > 0:
            n -= 1
            result = chr(65 + (n % 26)) + result
            n //= 26
        return result
    
    def _finalize_component_group(self, header, components, dept_id, current_program, 
                                   sheet_name, col_mapping, line_items_data, program_line_items,
                                   page_no, line_no, category_code=None, category_name=None):
        """Finalize a group of components (with or without TOTAL) as a single amendment"""
        if not components:
            return
        
        # Build description: header + components
        description_parts = [header] if header else []
        for comp in components:
            description_parts.append(comp['description'])
        combined_description = ' | '.join(description_parts)
        
        # Sum up amounts from components
        total_original = sum(c['original_gab'] for c in components)
        total_final = sum(c['final_amount'] for c in components)
        total_increase = sum(c['increase'] for c in components)
        total_decrease = abs(sum(c['decrease'] for c in components))
        total_net_change = total_final - total_original
        
        # Use stored page_no/line_no from first component
        amendment_page_no = components[0].get('page_no') if components and components[0].get('page_no') else page_no
        amendment_line_no = components[0].get('line_no') if components and components[0].get('line_no') else line_no
        
        # Use agency from first component (they should all be the same)
        final_category_code = components[0].get('agency_code') if components and components[0].get('agency_code') else category_code
        final_category_name = components[0].get('agency_name') if components and components[0].get('agency_name') else category_name
        
        # Determine amendment type
        if total_original == 0:
            amendment_type = 'CREATE'
        elif total_net_change > 0:
            amendment_type = 'INCREASE'
        elif total_net_change < 0:
            amendment_type = 'DECREASE'
        else:
            amendment_type = 'RETAIN'
        
        # Create the line item
        self.line_item_counter += 1
        final_dept_id = current_program.get('department_id', dept_id) if current_program else dept_id
        
        assignment_method = "sheet_match"
        assignment_reason = f"Assigned from sheet '{sheet_name}' using department matching logic"
        
        line_item = {
            "id": f"{final_dept_id}-LINE-{self.line_item_counter:04d}",
            "program_id": current_program['id'] if current_program else None,
            "department_id": final_dept_id,
            "page_no": amendment_page_no,
            "line_no": amendment_line_no,
            "description": combined_description[:500],
            "amendment_type": amendment_type,
            "original_amount": float(total_original),
            "final_amount": float(total_final),
            "increase": float(total_increase),
            "decrease": float(total_decrease),
            "net_change": float(total_net_change),
            "percent_change": float((total_net_change / total_original * 100) if total_original > 0 else 0.0),
            "formulas": components[0].get('row_formulas') if components else None,
            "excel_row": components[0]['row_idx'] if components else None,
            "excel_sheet": sheet_name,
            "source": {
                "file": "Annex A - Line By Line Amendments.xlsx",
                "sheet": sheet_name,
                "row": components[0]['row_idx'] if components else None,
                "cell_reference": f"{self.col_num_to_letter(col_mapping.get('line item', 2) + 1)}{components[0]['row_idx']}" if components and col_mapping.get('line item') is not None else f"A{components[0]['row_idx'] if components else ''}",
                "columns": {
                    "description": self.col_num_to_letter(col_mapping.get('line item', 2) + 1) if col_mapping.get('line item') is not None else "C",
                    "gab": self.col_num_to_letter(col_mapping.get('gab', 6) + 1) if col_mapping.get('gab') is not None else "G",
                    "committee": self.col_num_to_letter(col_mapping.get('committee report', 7) + 1) if col_mapping.get('committee report') is not None else "H",
                    "sv_committee": self.col_num_to_letter(col_mapping.get('sv committee', 10) + 1) if col_mapping.get('sv committee') is not None else "K"
                }
            },
            "assignment": {
                "method": assignment_method,
                "reason": assignment_reason,
                "original_sheet_dept": dept_id,
                "final_dept": final_dept_id
            },
                                    "is_grouped": True,
                                    "component_count": len(components),
                                    "header": header,
                                    "agency_code": final_category_code,
                                    "agency_name": final_category_name
                                }
        
        line_items_data.append(line_item)
        if current_program:
            program_line_items.append(line_item)
            current_program['original_amount'] += total_original
            current_program['final_amount'] += total_final
            current_program['increase'] += total_increase
            current_program['decrease'] += total_decrease
        
    def parse_amount(self, value) -> float:
        """Convert various amount formats to float"""
        if pd.isna(value) or value is None:
            return 0.0
        
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            # Remove currency symbols, commas, spaces
            cleaned = re.sub(r'[₱,\s]', '', value)
            try:
                return float(cleaned)
            except ValueError:
                return 0.0
        
        return 0.0
    
    def normalize_dept_name(self, name: str) -> str:
        """Normalize department name"""
        if not name:
            return "Unknown"
        
        # Common abbreviations
        mappings = {
            "DEPT": "Department",
            "DEP'T": "Department",
            "NATL": "National",
            "GOVT": "Government"
        }
        
        normalized = str(name).strip()
        for abbr, full in mappings.items():
            normalized = normalized.replace(abbr, full)
        
        return normalized
    
    def _generate_dept_code(self, name: str) -> str:
        """Generate department code from name"""
        if not name:
            return "UNK"
        
        # Common codes
        code_mappings = {
            "Department of Public Works and Highways": "DPWH",
            "Department of Education": "DepED",
            "Department of Agriculture": "DA",
            "Department of Health": "DOH",
            "Department of National Defense": "DND",
            "Department of Interior and Local Government": "DILG",
            "Department of Budget and Management": "DBM",
            "Office of the President": "OP",
            "Office of the Vice President": "OVP",
            "Congress of the Philippines": "Congress",
            "Judiciary": "Judiciary"
        }
        
        name_upper = str(name).upper()
        for full_name, code in code_mappings.items():
            if full_name.upper() in name_upper:
                return code
        
        # Generate code from initials
        words = str(name).split()
        if words:
            # Take first 3 significant words (skip common prefixes)
            significant_words = [w for w in words if w.upper() not in ['THE', 'OF', 'AND', 'FOR', 'A', 'AN']][:3]
            if significant_words:
                code = ''.join([w[0].upper() for w in significant_words if w])
                return code if code else "UNK"
        
        return "UNK"
    
    def _match_sheet_to_department(self, sheet_name: str) -> str:
        """Match sheet name to department ID from parsed departments"""
        if not self.departments:
            return self._generate_dept_code(sheet_name)
        
        sheet_upper = sheet_name.upper().strip()
        sheet_clean = sheet_upper.replace('-', ' ').replace('_', ' ')
        
        # Sheet name to department mapping
        # Most sheets starting with 'D' are departments
        # SUCs-* sheets are under DepEd (AOO-5)
        # DOTR = Transportation, DOT = Tourism
        sheet_to_dept_mapping = {
            'CONGRESS': ['CONGRESS', 'HOUSE', 'SENATE'],
            'OP': ['PRESIDENT', 'OFFICE OF THE PRESIDENT', 'EXECUTIVE SECRETARY'],
            'OEOs': ['PRESIDENT', 'OFFICE OF THE PRESIDENT', 'EXECUTIVE SECRETARY', 'MOVIE', 'TELEVISION', 'REVIEW', 'CLASSIFICATION', 'MTRCB', 'MTVRCB'],
            'OVP': ['VICE PRESIDENT', 'OFFICE OF THE VICE PRESIDENT'],
            'DAR': ['AGRARIAN REFORM'],
            'DA': ['AGRICULTURE'],
            'DEPDEV': ['ECONOMY', 'PLANNING', 'DEVELOPMENT'],
            'DEPED': ['EDUCATION'],
            'SUCS-UP': ['EDUCATION', 'DEPED'],  # All SUCs-* are under DepEd
            'SUCS-NCR': ['EDUCATION', 'DEPED'],
            'SUCS-RI': ['EDUCATION', 'DEPED'],
            'SUCS-CAR': ['EDUCATION', 'DEPED'],
            'SUCS-RII': ['EDUCATION', 'DEPED'],
            'SUCS-RIII': ['EDUCATION', 'DEPED'],
            'SUCS-RIV-A': ['EDUCATION', 'DEPED'],
            'SUCS-RIV-B': ['EDUCATION', 'DEPED'],
            'SUCS-RV': ['EDUCATION', 'DEPED'],
            'SUCS-RVI': ['EDUCATION', 'DEPED'],
            'SUCS-NIR': ['EDUCATION', 'DEPED'],
            'SUCS-RVII': ['EDUCATION', 'DEPED'],
            'SUCS-RVIII': ['EDUCATION', 'DEPED'],
            'SUCS-RIX': ['EDUCATION', 'DEPED'],
            'SUCS-RX': ['EDUCATION', 'DEPED'],
            'SUCS-RXI': ['EDUCATION', 'DEPED'],
            'SUCS-RXII': ['EDUCATION', 'DEPED'],
            'SUCS-RXIII': ['EDUCATION', 'DEPED'],
            'SUCS-BARMM': ['EDUCATION', 'DEPED'],
            'DENR': ['ENVIRONMENT', 'NATURAL RESOURCES'],
            'DOF': ['FINANCE'],
            'DFA': ['FOREIGN AFFAIRS'],
            'DOH': ['HEALTH'],
            'DHSUD': ['HOUSING', 'SETTLEMENT'],
            'DICT': ['INFORMATION', 'COMMUNICATIONS', 'TECHNOLOGY'],
            'DILG': ['INTERIOR', 'LOCAL GOVERNMENT'],
            'DOJ': ['JUSTICE'],
            'DOLE': ['LABOR', 'EMPLOYMENT'],
            'DMW': ['MIGRANT WORKERS'],
            'DND': ['DEFENSE', 'NATIONAL DEFENSE'],
            'DPWH': ['PUBLIC WORKS', 'HIGHWAYS'],
            'DOST': ['SCIENCE', 'TECHNOLOGY'],
            'DSWD': ['SOCIAL WELFARE', 'DEVELOPMENT'],
            'DOT': ['TOURISM'],  # Department of Tourism (NOT Transportation)
            'DTI': ['TRADE', 'INDUSTRY'],
            'DOTR': ['OFFICE OF THE SECRETARY.*TRANSPORTATION', 'AOS-17'],  # Department of Transportation
            'JUDICIARY': ['JUDICIARY', 'SUPREME COURT'],
            'COA': ['COMMISSION ON AUDIT'],
            'BSGC': ['BUDGETARY SUPPORT', 'GOVERNMENT CORPORATIONS', 'GOCC'],
            'ALGU': ['ALLOCATION', 'LGU', 'LOCAL GOVERNMENT UNIT'],
            'MPBF': [],  # As-is for now
            'NDRRMF': ['DISASTER', 'RISK', 'REDUCTION', 'MANAGEMENT', 'FUND'],
            'RAFPMP': [],  # As-is for now
            'UA': ['UNPROGRAMMED', 'APPROPRIATIONS'],
            'DOE-ATTACHED CORP': ['ENERGY', 'GOCC'],  # GOCCs under Department of Energy
            'NIA': ['IRRIGATION'],
        }
        
        # Check if sheet starts with 'SUCS-' (all are under DepEd - AOO-5)
        if sheet_upper.startswith('SUCS-'):
            for dept in self.departments:
                dept_id = dept.get('id', '').upper()
                dept_name = dept.get('name', '').upper()
                # Look for DepEd Office of the Secretary (AOO-5)
                if dept_id == 'AOO-5' or ('EDUCATION' in dept_name and 'OFFICE OF THE SECRETARY' in dept_name):
                    return dept['id']
        
        # Try direct mapping first
        for sheet_key, dept_keywords in sheet_to_dept_mapping.items():
            if sheet_key in sheet_upper or sheet_upper in sheet_key:
                # Find department matching keywords
                # For DOTR, prioritize AOS-17 (Office of the Secretary - Transportation)
                if sheet_key == 'DOTR':
                    for dept in self.departments:
                        dept_id = dept.get('id', '').upper()
                        dept_name = dept.get('name', '').upper()
                        # First check for exact AOS-17 match
                        if dept_id == 'AOS-17':
                            return dept['id']
                    # Then check for Office of Secretary + Transportation (must have both)
                    for dept in self.departments:
                        dept_name = dept.get('name', '').upper()
                        if 'OFFICE OF THE SECRETARY' in dept_name and 'TRANSPORTATION' in dept_name:
                            return dept['id']
                
                # For other sheets, use keyword matching
                for dept in self.departments:
                    dept_name = dept.get('name', '').upper()
                    dept_id = dept.get('id', '').upper()
                    dept_code = dept.get('code', '').upper()
                    # Check if any keyword matches department name, ID, or code
                    if any(keyword in dept_name for keyword in dept_keywords) or \
                       any(keyword in dept_id for keyword in dept_keywords) or \
                       any(keyword in dept_code for keyword in dept_keywords):
                        return dept['id']
        
        # Try to find matching department - multiple strategies
        best_match = None
        best_score = 0
        
        for dept in self.departments:
            dept_name = dept.get('name', '').upper()
            dept_code = dept.get('code', '').upper()
            dept_id = dept.get('id', '').upper()
            
            score = 0
            
            # Exact matches get highest score
            if sheet_upper == dept_code or sheet_upper == dept_id:
                return dept['id']  # Perfect match, return immediately
            
            # Check if sheet name contains department code/ID or vice versa
            if dept_code and dept_code in sheet_upper:
                score += 10
            if dept_id and dept_id in sheet_upper:
                score += 10
            if sheet_upper in dept_code or sheet_upper in dept_id:
                score += 10
            
            # Check for word matches in department name
            dept_words = set(word for word in dept_name.split() if len(word) > 2)
            sheet_words = set(word for word in sheet_clean.split() if len(word) > 2)
            common_words = dept_words.intersection(sheet_words)
            if common_words:
                score += len(common_words) * 3
            
            # Check if department name contains sheet name or vice versa
            if sheet_upper in dept_name:
                score += 5
            if any(word in dept_name for word in sheet_words if len(word) > 3):
                score += 3
            
            if score > best_score:
                best_score = score
                best_match = dept['id']
        
        # Return best match if score is good enough, otherwise generate code
        if best_score >= 5:
            return best_match
        
        # Fallback to generating code
        return self._generate_dept_code(sheet_name)
    
    def parse_general_summary(self) -> Dict:
        """Parse General Summary file for department-level totals - Hybrid approach (Values + Formulas)"""
        file_path = self.base_dir / "General Summary of Committee Report No. 18 on House Bill No. 4058 (FY 2026 GAB).xlsx"
        
        print(f"📊 Parsing General Summary...")
        
        try:
            import openpyxl
            
            # PASS 1: Read VALUES (data_only=True)
            wb_values = openpyxl.load_workbook(file_path, data_only=True)
            ws_values = wb_values['By Agency']
            
            # PASS 2: Read FORMULAS (data_only=False)
            wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
            ws_formulas = wb_formulas['By Agency']
            
            dept_data = []
            current_parent_dept = None
            
            # Iterate through all rows
            for row_num in range(1, ws_values.max_row + 1):
                # Get value row and formula row
                row_val = ws_values[row_num]
                row_form = ws_formulas[row_num]
                
                # Get values from columns
                col_a = row_val[0].value
                col_b = row_val[1].value
                col_c = row_val[2].value
                col_d = row_val[3].value
                col_e = row_val[4].value
                col_f = row_val[5].value
                col_g = row_val[6].value
                col_h = row_val[7].value
                col_i = row_val[8].value
                
                # DEBUG: Print first few rows to check structure
                if row_num <= 10:
                    print(f"   Row {row_num}: A={col_a}, B={col_b}, C={col_c}, E={col_e}")

                # Track parent department (top-level departments have roman numerals in column A)
                # Structure: Column A = "I.", "II.", etc., Column B = "DEPARTMENT OF X" -> Top-level department
                if col_a and col_b and not col_c:
                    col_a_str = str(col_a).strip()
                    col_b_str = str(col_b).strip() if col_b else ""
                    # Check if column A is a roman numeral (I., II., III., etc.)
                    is_roman_numeral = (
                        col_a_str.endswith('.') and col_a_str[:-1] in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII', 'XIII', 'XLIII'] or
                        col_a_str in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII', 'XIII', 'XLIII']
                    )
                    if is_roman_numeral and len(col_b_str) > 5:
                        current_parent_dept = col_b_str
                
                # Build full department name
                dept_name = None
                is_aggregate = False
                parent_department = current_parent_dept
                prefix = None
                prefix_is_letter_code = False
                
                # Check for top-level department: Column A has roman numeral, Column B has department name
                # This MUST be checked first before agencies
                is_top_level_dept = False
                if col_a and col_b and not col_c:
                    col_a_str = str(col_a).strip()
                    col_b_str = str(col_b).strip()
                    is_roman_numeral = (
                        (col_a_str.endswith('.') and col_a_str[:-1] in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII', 'XIII', 'XLIII']) or
                        col_a_str in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'XI', 'XII', 'XIII', 'XLIII']
                    )
                    if is_roman_numeral and len(col_b_str) > 5:
                        # This is a top-level department
                        dept_name = col_b_str
                        prefix = col_a_str
                        # Not an agency - this is a department
                        prefix_is_letter_code = False
                        is_top_level_dept = True
                
                # Check for agency: Column A empty, Column B has letter code (A., B., etc.), Column C has name
                # Only process if we haven't already identified a top-level department
                if not is_top_level_dept and not col_a and col_b and col_c:
                    prefix = str(col_b).strip()
                    name = str(col_c).strip()
                    
                    if len(name) < 5 or ('department' in name.lower() and 'agency' not in name.lower()):
                        continue
                    
                    name_lower = name.lower()
                    is_aggregate = any(keyword in name_lower for keyword in [
                        'total', 'subtotal', 'sum', 'aggregate', 'all', 'combined', 'grand total'
                    ])
                    
                    # Check if prefix is a letter code (A., B., C., AA., AB., etc.) or sub-agency code (G.1, J.2, etc.)
                    # This indicates an agency, not a top-level department
                    # Letter codes can be single (A., B., C.) or double (AA., AB., AC., etc.)
                    prefix_is_letter_code = (
                        prefix and (
                            (len(prefix) == 2 and prefix[0].isalpha() and prefix[1] == '.') or  # A., B., C., etc.
                            (len(prefix) == 3 and prefix[0].isalpha() and prefix[1].isalpha() and prefix[2] == '.') or  # AA., AB., AC., etc.
                            (len(prefix) >= 4 and prefix[0].isalpha() and prefix[1] == '.' and prefix[2:].isdigit())  # G.1, J.2, etc.
                        )
                    )
                    
                    if 'office' in name_lower and 'secretary' in name_lower and parent_department:
                        dept_name = f"{prefix} {name} ({parent_department})"
                    elif col_d and str(col_d).strip():
                        dept_name = f"{prefix} {name} - {str(col_d).strip()}"
                    else:
                        dept_name = f"{prefix} {name}"
                
                elif col_a and col_c and not col_b:
                    col_a_str = str(col_a).strip()
                    if col_a_str and (col_a_str.endswith('.') or col_a_str in ['I', 'II', 'III', 'IV', 'V', 'VI']):
                        name = str(col_c).strip()
                        if len(name) > 5:
                            dept_name = f"{col_a_str} {name}"
                
                if not dept_name:
                    continue
                
                # Extract budget amounts (VALUES)
                original_amount = (float(col_e) * 1000) if col_e and isinstance(col_e, (int, float)) else 0.0
                increase = (float(col_f) * 1000) if col_f and isinstance(col_f, (int, float)) else 0.0
                decrease = (float(col_g) * 1000) if col_g and isinstance(col_g, (int, float)) else 0.0
                net_change = (float(col_h) * 1000) if col_h and isinstance(col_h, (int, float)) else 0.0
                final_amount = (float(col_i) * 1000) if col_i and isinstance(col_i, (int, float)) else 0.0
                
                # Extract FORMULAS
                formulas = {}
                if row_form[4].data_type == 'f': formulas['original_amount'] = row_form[4].value
                if row_form[5].data_type == 'f': formulas['increase'] = row_form[5].value
                if row_form[6].data_type == 'f': formulas['decrease'] = row_form[6].value
                if row_form[7].data_type == 'f': formulas['net_change'] = row_form[7].value
                if row_form[8].data_type == 'f': formulas['final_amount'] = row_form[8].value
                
                # Calculate derived values if needed
                if net_change == 0 and (increase != 0 or decrease != 0):
                    net_change = increase + decrease
                if final_amount == 0 and original_amount > 0:
                    final_amount = original_amount + net_change
                
                if original_amount == 0 and final_amount == 0:
                    continue
                
                self.dept_counter += 1
                base_code = self._generate_dept_code(dept_name)
                dept_code = base_code
                
                existing_codes = [d['code'] for d in dept_data]
                if dept_code in existing_codes:
                    suffix = 1
                    while f"{dept_code}-{suffix}" in existing_codes:
                        suffix += 1
                    dept_code = f"{dept_code}-{suffix}"
                
                # Determine if this is a department or an agency
                # Top-level departments have roman numerals in column A - they are NOT agencies
                # Agencies have patterns like:
                # - "A. Office of the Secretary (DEPARTMENT OF X)"
                # - Start with letter codes like "A.", "B.", "C.", etc. AND have a parent in parentheses
                # - "A. Senate", "B. Senate Electoral Tribunal" are agencies under Congress
                # - "A. The President's Offices" is an agency under OP
                # - "G. Armed Forces of the Philippines" is an agency under DND
                # - "J. Philippine National Police" is an agency under DILG
                # - Any entry starting with letter code (including sub-codes like G.1, J.2) that appears after a parent department is an agency
                is_agency = False
                
                # If this is a top-level department (roman numeral), it's NOT an agency
                if is_top_level_dept:
                    is_agency = False
                # Check if prefix is a letter code (indicates agency)
                elif prefix_is_letter_code:
                    # If it has a parent department in parentheses, it's an agency
                    if '(' in dept_name and ')' in dept_name and 'DEPARTMENT' in dept_name:
                        is_agency = True
                    # If there's a current_parent_dept set, this is an agency under that department
                    elif current_parent_dept:
                        is_agency = True
                    # If it's under a known parent (Congress, OP, OVP), it's an agency
                    elif any(parent in dept_name for parent in ['Senate', 'House', 'President', 'Vice-President', 'Vice President']):
                        is_agency = True
                    # If it's "Office of the Secretary" with a department in parentheses
                    elif 'Office of the Secretary' in dept_name and '(' in dept_name:
                        is_agency = True
                    # Known agency patterns
                    elif 'Armed Forces' in dept_name or 'Armed Forces of the Philippines' in dept_name:
                        is_agency = True
                    elif 'Philippine National Police' in dept_name or 'PNP' in dept_name:
                        is_agency = True
                    # If prefix is a letter code, it's likely an agency (unless it's a top-level department)
                    else:
                        is_agency = True
                
                # Also check if dept_name starts with letter code (fallback) - including double letters (AA., AB., etc.)
                # Check for single letter codes
                starts_with_letter = dept_name.startswith(('A.', 'B.', 'C.', 'D.', 'E.', 'F.', 'G.', 'H.', 'I.', 'J.', 'K.', 'L.', 'M.', 'N.', 'O.', 'P.', 'Q.', 'R.', 'S.', 'T.', 'U.', 'V.', 'W.', 'X.', 'Y.', 'Z.'))
                # Check for double letter codes (AA., AB., AC., etc.)
                starts_with_double_letter = re.match(r'^[A-Z]{2}\.', dept_name)
                if (starts_with_letter or starts_with_double_letter) and not is_agency:
                    # Check for sub-agency codes (G.1, J.2, etc.)
                    if re.match(r'^[A-Z]\.\d+', dept_name):
                        is_agency = True
                    # If there's a current_parent_dept, it's an agency
                    elif current_parent_dept:
                        is_agency = True
                
                # Extract parent department name if it's an agency
                parent_dept_name = None
                if is_agency:
                    if '(' in dept_name and ')' in dept_name:
                        # Extract parent from parentheses: "A. Office of the Secretary (DEPARTMENT OF EDUCATION)"
                        parent_match = dept_name.split('(')[-1].split(')')[0].strip()
                        if 'DEPARTMENT' in parent_match.upper():
                            parent_dept_name = parent_match
                    elif current_parent_dept:
                        # Use the tracked parent department
                        parent_dept_name = current_parent_dept
                    elif 'Armed Forces' in dept_name or 'Armed Forces of the Philippines' in dept_name:
                        # Armed Forces is under Department of National Defense
                        parent_dept_name = "DEPARTMENT OF NATIONAL DEFENSE"
                
                # Clean department/agency name - remove numbering prefixes
                clean_name = dept_name
                # Remove roman numeral prefixes (I., II., III., IV., V., etc.)
                clean_name = re.sub(r'^[IVXLCDM]+\.\s*', '', clean_name, flags=re.IGNORECASE)
                # Remove numeric prefixes (1., 2., 3., etc.) - for both departments and agencies
                clean_name = re.sub(r'^\d+\.\s*', '', clean_name)
                # Remove letter prefixes (A., B., C., AA., AB., etc.) - but only if it's an agency
                if is_agency:
                    # Remove single letter prefixes (A., B., C., etc.)
                    clean_name = re.sub(r'^[A-Z]\.\s*', '', clean_name)
                    # Remove double letter prefixes (AA., AB., AC., etc.)
                    clean_name = re.sub(r'^[A-Z]{2}\.\s*', '', clean_name)
                
                dept_entry = {
                    "id": dept_code,
                    "code": dept_code,
                    "name": self.normalize_dept_name(clean_name),
                    "original_amount": original_amount,
                    "increase": increase,
                    "decrease": decrease,
                    "net_change": net_change,
                    "final_amount": final_amount,
                    "percent_change": (net_change / original_amount * 100) if original_amount > 0 else 0.0,
                    "amendment_count": 0,
                    "program_count": 0,
                    "project_count": 0,
                    "source_sheets": ["General Summary"],
                    "excel_row": row_num,
                    "is_aggregate": is_aggregate,
                    "parent_department": parent_department if parent_department else None,
                    "is_agency": is_agency,  # Flag to distinguish agencies from departments
                    "parent_department_name": parent_dept_name,  # Parent department name for agencies
                    "formulas": formulas if formulas else None,
                    # Enhanced traceability fields
                    "source": {
                        "file": "General Summary of Committee Report No. 18 on House Bill No. 4058 (FY 2026 GAB).xlsx",
                        "sheet": "By Agency",
                        "row": row_num,
                        "cell_reference": f"C{row_num}",  # Column C is department name
                        "columns": {
                            "prefix": "B",  # Column B
                            "name": "C",    # Column C
                            "original": "E",  # Column E
                            "increase": "F",  # Column F
                            "decrease": "G",  # Column G
                            "net_change": "H",  # Column H
                            "final": "I"  # Column I
                        }
                    },
                    "assignment": {
                        "method": "general_summary",
                        "reason": f"Parsed from General Summary row {row_num}, column structure: prefix in B, name in C, amounts in E-I"
                    }
                }
                
                dept_data.append(dept_entry)
            
            wb_values.close()
            wb_formulas.close()
            
            print(f"✅ Parsed {len(dept_data)} departments from General Summary")
            return dept_data
            
        except Exception as e:
            print(f"❌ Error parsing General Summary: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def parse_annex_a(self) -> List[Dict]:
        """
        Parse Annex A - Line By Line Amendments (53 sheets) - Hybrid approach (Values + Formulas)
        
        ALGORITHM FOR READING ANNEX A:
        ==============================
        1. For each worksheet (53 total):
           a. Find header row (contains "Page", "Line", "GAB", "Committee")
           b. Read rows starting from data_start_row
           
        2. For each row:
           a. Column A: 
              - If matches regex ^[A-Z]\.$ (letter followed by dot): This is an AGENCY CODE
              - Store agency code (A., B., C., G., J., etc.)
              - Get agency name from Column C (Column B is Line No)
              - All subsequent rows belong to this agency until a new agency code is read
              - If not an agency code, might be a page number
              
           b. Column B: Line number (retain in memory)
           
           c. Columns C-F: Description (Line Item of Appropriations, Implementing/Operating Unit, Object of Expenditures)
           
           d. Columns G-K: Amounts
              - G = GAB (original)
              - H = Increase
              - I = Decrease
              - J = Net Change
              - K = Committee Report (final)
              
        3. Processing logic:
           - If "By Activity/Project:" or "By Object of Expenditure:" found: Set grouping type
           - If header row (no G-K values, has description): This is a program/activity name
           - If component row (has G-K values, not TOTAL, has current_header): Add to pending_components
           - If TOTAL row: Finalize component group
           - If standalone line item (has G-K values, no current_header): Process as individual amendment
           
        4. Agency tracking:
           - When new agency code (A., B., etc.) is read in Column A, switch to that agency
           - All line items processed after agency code belong to that agency
           - Agency code and name are stored and used for all subsequent rows until next agency code
        """
        file_path = self.base_dir / "Annex A - Line By Line Amendments.xlsx"
        
        print(f"📋 Parsing Annex A (Line By Line Amendments)...")
        
        try:
            import openpyxl
            
            # PASS 1: Read VALUES (data_only=True)
            print("   Reading values...")
            wb_values = openpyxl.load_workbook(file_path, data_only=True)
            
            # PASS 2: Read FORMULAS (data_only=False)
            print("   Reading formulas...")
            wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
            
            print(f"   Found {len(wb_values.sheetnames)} sheets")
            
            programs_data = []
            line_items_data = []
            annex_a_departments = []  # Department-level entries from Annex A worksheets
            
            for sheet_name in wb_values.sheetnames:
                print(f"   Processing sheet: {sheet_name}")
                
                try:
                    ws_val = wb_values[sheet_name]
                    ws_form = wb_formulas[sheet_name]
                    
                    dept_id = self._match_sheet_to_department(sheet_name) if self.departments else self._generate_dept_code(sheet_name)
                    
                    # Special handling for DOTR - must match AOS-17
                    if sheet_name.upper() == 'DOTR':
                        aos17 = next((d for d in self.departments if d.get('id', '').upper() == 'AOS-17'), None)
                        if aos17:
                            dept_id = aos17['id']
                            print(f"      ✓ DOTR sheet matched to {dept_id}: {aos17.get('name', '')[:50]}")
                    
                    # Find header row - rows 5-6 contain column descriptions
                    # Row 5: "Page No. | Line No. | Line Item of Appropriations | ... | FY 2026 GAB | COMMITTEE REPORT | ... | FY 2026 SV Committee"
                    # Row 6: " |  |  |  |  |  | (HBN 4058) | Increase | Decrease | Net Change |"
                    header_row = None
                    # Check rows 5-6 first (standard location)
                    for row_idx in [5, 6]:
                        if row_idx <= ws_val.max_row:
                            row_values = [str(ws_val.cell(row_idx, col).value or '').strip() for col in range(1, min(25, ws_val.max_column + 1))]
                            row_str = ' '.join(row_values).lower()
                            if 'page' in row_str and 'line' in row_str:
                                header_row = row_idx
                                break
                    
                    # Fallback: search other rows if 5-6 don't have it
                    if header_row is None:
                        for row_idx in range(1, min(30, ws_val.max_row + 1)):
                            if row_idx in [5, 6]:  # Skip, already checked
                                continue
                            row_values = [str(ws_val.cell(row_idx, col).value or '').strip() for col in range(1, min(25, ws_val.max_column + 1))]
                            row_str = ' '.join(row_values).lower()
                            if 'page' in row_str and 'line' in row_str:
                                header_row = row_idx
                                break
                    
                    if header_row is None:
                        print(f"      ⚠️  Could not find header row in {sheet_name}")
                        continue
                    
                    # Use both row 5 and row 6 for column mapping (they contain split headers)
                    header_row_5 = 5 if ws_val.max_row >= 5 else None
                    header_row_6 = 6 if ws_val.max_row >= 6 else None
                    
                    # Extract worksheet description from row 10 (remove roman numeral prefix)
                    # Row 10 contains the department-level name (e.g., "I. CONGRESS OF THE PHILIPPINES" -> "CONGRESS OF THE PHILIPPINES")
                    worksheet_description = sheet_name  # Default to sheet name
                    dept_name_from_row10 = None
                    if ws_val.max_row >= 10:
                        row_10_val = ws_val.cell(10, 1).value
                        if row_10_val:
                            desc_str = str(row_10_val).strip()
                            # Remove roman numeral prefix (I., II., III., IV., V., VI., VII., VIII., IX., X., XI., XII., XIII., XXI., XVII., etc.)
                            # Match patterns like "I. ", "II. ", "XXI. ", "XVII. " etc.
                            desc_str = re.sub(r'^[IVXLCDM]+\.\s*', '', desc_str, flags=re.IGNORECASE)
                            desc_str = re.sub(r'^[IVXLCDM]+\s+', '', desc_str, flags=re.IGNORECASE)
                            # Also handle cases without dot: "I ", "II ", etc.
                            desc_str = re.sub(r'^[IVXLCDM]+\s+', '', desc_str, flags=re.IGNORECASE)
                            if desc_str and len(desc_str) > 3:
                                worksheet_description = desc_str.strip()
                                dept_name_from_row10 = desc_str.strip()
                    
                    # Create department-level entry from this worksheet
                    # Use row 10 department name, or fall back to sheet name
                    annex_dept_name = dept_name_from_row10 if dept_name_from_row10 else worksheet_description
                    annex_dept_code = self._generate_dept_code(annex_dept_name)
                    
                    # Check if this department already exists (from General Summary or previous worksheet)
                    existing_dept = next((d for d in annex_a_departments if d.get('name', '').upper() == annex_dept_name.upper()), None)
                    if not existing_dept:
                        # Create new department entry from Annex A worksheet
                        annex_a_departments.append({
                            "id": annex_dept_code,
                            "code": annex_dept_code,
                            "name": self.normalize_dept_name(annex_dept_name),
                            "sheet_name": sheet_name,
                            "original_amount": 0.0,
                            "increase": 0.0,
                            "decrease": 0.0,
                            "net_change": 0.0,
                            "final_amount": 0.0,
                            "percent_change": 0.0,
                            "amendment_count": 0,
                            "program_count": 0,
                            "project_count": 0,
                            "source": "Annex A",
                            "is_agency": False
                        })
                    
                    # Check if this is a SUCs sheet (huge component under DepEd)
                    is_sucs_sheet = sheet_name.upper().startswith('SUCS-')
                    
                    # Map columns
                    col_mapping = {
                        'page': None, 'line': None, 'line item': None,
                        'gab': None, 'committee report': None, 'sv committee': None,
                        'increase': None, 'decrease': None, 'net change': None,
                        'category': None  # Column A for category/subdivision (A., B., C., etc.)
                    }
                    
                    # Map columns using both row 5 and row 6 (headers are split across two rows)
                    # Column A is always the category/subdivision (A., B., C., etc.) or page number
                    col_mapping['category'] = 0  # Column A (index 0)
                    
                    # Check row 5 first
                    if header_row_5:
                        header_cells_5 = list(ws_val[header_row_5])
                        for col_idx, cell in enumerate(header_cells_5):
                            if cell.value:
                                cell_val = str(cell.value).strip()
                                cell_val_lower = cell_val.lower().replace('\n', ' ').replace('\r', ' ')
                                
                                if 'page' in cell_val_lower and col_mapping['page'] is None: col_mapping['page'] = col_idx
                                if 'line no' in cell_val_lower and col_mapping['line'] is None: col_mapping['line'] = col_idx
                                if 'line item' in cell_val_lower and col_mapping['line item'] is None: col_mapping['line item'] = col_idx
                                if ('fy 2026 gab' in cell_val_lower or 'gab' in cell_val_lower) and col_mapping['gab'] is None: col_mapping['gab'] = col_idx
                                if 'committee report' in cell_val_lower and 'sv' not in cell_val_lower and col_mapping['committee report'] is None: col_mapping['committee report'] = col_idx
                                if ('sv committee' in cell_val_lower or 'sv committee report' in cell_val_lower) and col_mapping['sv committee'] is None: col_mapping['sv committee'] = col_idx
                    
                    # Check row 6 for additional headers (Increase, Decrease, Net Change)
                    if header_row_6:
                        header_cells_6 = list(ws_val[header_row_6])
                        for col_idx, cell in enumerate(header_cells_6):
                            if cell.value:
                                cell_val = str(cell.value).strip()
                                cell_val_lower = cell_val.lower().replace('\n', ' ').replace('\r', ' ')
                                
                                if 'increase' in cell_val_lower and col_mapping['increase'] is None: col_mapping['increase'] = col_idx
                                if 'decrease' in cell_val_lower and col_mapping['decrease'] is None: col_mapping['decrease'] = col_idx
                                if 'net change' in cell_val_lower and col_mapping['net change'] is None: col_mapping['net change'] = col_idx
                                if ('hbn 4058' in cell_val_lower or 'committee report' in cell_val_lower) and col_mapping['committee report'] is None: col_mapping['committee report'] = col_idx
                    
                    # Find data start row
                    data_start_row = header_row + 1
                    for check_row in range(header_row + 1, min(header_row + 10, ws_val.max_row + 1)):
                        has_amount = False
                        has_page_line = False
                        
                        for col_idx in [6, 7, 10]:
                            if col_idx < ws_val.max_column:
                                cell_val = ws_val.cell(check_row, col_idx + 1).value
                                if isinstance(cell_val, (int, float)) and cell_val > 0:
                                    has_amount = True
                                break
                        
                        page_val = ws_val.cell(check_row, col_mapping['page'] + 1).value if col_mapping['page'] is not None else None
                        if page_val and str(page_val).strip() and str(page_val).strip().lower() not in ['none', 'note:']:
                            has_page_line = True
                            
                        desc_cell = ws_val.cell(check_row, col_mapping['line item'] + 1).value if col_mapping['line item'] is not None else None
                        desc_text = str(desc_cell or '').lower()
                        is_note = 'note:' in desc_text or desc_text.startswith('note') or 'total' in desc_text
                        
                        if (has_amount or has_page_line) and not is_note:
                            data_start_row = check_row
                            break
                    
                    # Process rows
                    current_program = None
                    program_line_items = []
                    pending_multiline_delete = None  # Track multi-line DELETE operations
                    
                    # Track values in memory (retain across rows)
                    stored_page_no = None  # Retain until new page number is read
                    stored_line_no = None  # Retain until new line number is read
                    stored_agency_code = None  # A., B., C., etc. - switches when new agency code is read
                    stored_agency_name = None  # Full agency name
                    
                    # Track grouping indicators and program/activity structure
                    current_grouping_type = None  # "By Activity/Project:", "By Object of Expenditure:", etc.
                    current_activity_name = None  # Program/Activity name (e.g., "General Management and Supervision")
                    current_activity_page = None  # Page number for the activity
                    current_activity_line = None  # Line number for the activity
                    pending_components = []  # Components (rows with G-K values) waiting for TOTAL
                    current_header = None  # Parent description header (C-F without G-K values)
                    in_insertion_context = False  # Track if we're processing items after "Between lines X and Y, insert..."
                    
                    for row_idx in range(data_start_row, ws_val.max_row + 1):
                        row_val = ws_val[row_idx]
                        row_form = ws_form[row_idx]
                        
                        # Read row by row and retain values
                        # Column A: Page No OR Agency header (A., B., C., G., J., etc.)
                        # Algorithm: If column A matches regex pattern [A-Z]\. (letter followed by dot), it's an agency code
                        # All subsequent rows belong to that agency until a new agency code is read
                        col_a_val = ws_val.cell(row_idx, 1).value if ws_val.max_column >= 1 else None
                        col_a_str = str(col_a_val).strip() if col_a_val else ''
                        
                        # Check if column A is an agency header using regex: letter followed by dot (A., B., C., G., J., etc.)
                        # This is the primary way to detect agencies in Annex A
                        is_agency_header = bool(re.match(r'^[A-Z]\.$', col_a_str))
                        
                        if is_agency_header:
                            # If we have pending components from previous agency, finalize them first
                            if pending_components and current_header:
                                self._finalize_component_group(
                                    current_header, pending_components, dept_id, current_program,
                                    sheet_name, col_mapping, line_items_data, program_line_items,
                                    page_no, line_no, category_code, category_name
                                )
                                pending_components = []
                                current_header = None
                                current_activity_name = None
                                current_activity_page = None
                                current_activity_line = None
                            
                            # Retain agency code
                            stored_agency_code = col_a_str
                            # Get agency name from column C (column B is Line No)
                            if ws_val.max_column >= 3:
                                col_c_val = ws_val.cell(row_idx, 3).value
                                if col_c_val and str(col_c_val).strip():
                                    agency_name = str(col_c_val).strip()
                                    if len(agency_name) > 3:
                                        stored_agency_name = agency_name
                            continue  # Skip processing agency header row as a line item
                        else:
                            # Column A might be a page number
                            if col_a_str:
                                try:
                                    page_num = float(col_a_str)
                                    if 0 < page_num < 10000:
                                        stored_page_no = str(int(page_num))
                                except:
                                    pass
                        
                        # Column B: Line No (purely for line numbers)
                        col_b_val = ws_val.cell(row_idx, 2).value if ws_val.max_column >= 2 else None
                        if col_b_val:
                            col_b_str = str(col_b_val).strip()
                            try:
                                line_num = float(col_b_str)
                                if 0 < line_num < 100000:
                                    stored_line_no = str(int(line_num))
                            except:
                                pass
                        
                        # Columns C-F: Description (Line Item of Appropriations, Implementing/Operating Unit, Object of Expenditures)
                        description_parts = []
                        for col_idx in range(3, min(7, ws_val.max_column + 1)):  # Columns C-F
                            cell_val = ws_val.cell(row_idx, col_idx).value
                            if cell_val and str(cell_val).strip():
                                description_parts.append(str(cell_val).strip())
                        line_item_text = ' | '.join(description_parts) if description_parts else ''
                        line_item_text_lower = line_item_text.lower()
                        
                        # Check for grouping indicators: "By Activity/Project:", "By Object of Expenditure:"
                        is_grouping_indicator = False
                        if 'by activity' in line_item_text_lower and ('project' in line_item_text_lower or '/' in line_item_text):
                            current_grouping_type = "By Activity/Project:"
                            is_grouping_indicator = True
                        elif 'by object' in line_item_text_lower and 'expenditure' in line_item_text_lower:
                            current_grouping_type = "By Object of Expenditure:"
                            is_grouping_indicator = True
                        
                        # Check for insertion description rows (hints, not actual line items)
                        # Examples: "Between lines 7 and 8, insert the following objects of expenditure:"
                        # These are just descriptions - skip them, but mark following items as CREATE
                        is_insertion_description = (
                            ('between lines' in line_item_text_lower and 'insert' in line_item_text_lower) or
                            ('after line' in line_item_text_lower and 'insert' in line_item_text_lower) or
                            ('before line' in line_item_text_lower and 'insert' in line_item_text_lower)
                        ) and not has_gk_values  # Only if it doesn't have amounts (it's just a description)
                        
                        if is_insertion_description:
                            # Check if this row also has pipe-separated entries after the insertion description
                            # Example: "Between lines 7 and 8, insert the following object of expenditures: | FINANCIAL ASSISTANCE/SUBSIDY | Communication Expenses"
                            if '|' in line_item_text:
                                # Extract the pipe-separated parts (remove the "Between lines..." prefix)
                                # Find where the pipe-separated entries start
                                pipe_parts = line_item_text.split('|')
                                if len(pipe_parts) > 1:
                                    # The first part is the insertion description, skip it
                                    # Process the remaining parts as separate entries
                                    entries = [p.strip() for p in pipe_parts[1:] if p.strip() and len(p.strip()) > 2]
                                    
                                    if entries and has_gk_values:
                                        # Create separate line items for each entry
                                        for entry_text in entries:
                                            self.line_item_counter += 1
                                            final_dept_id = current_program.get('department_id', dept_id) if current_program else dept_id
                                            
                                            # Split amounts equally among entries
                                            entry_original = original_gab / len(entries) if original_gab > 0 else 0.0
                                            entry_final = final_amount / len(entries) if final_amount > 0 else 0.0
                                            entry_increase = increase / len(entries) if increase > 0 else 0.0
                                            entry_decrease = decrease / len(entries) if decrease < 0 else 0.0
                                            entry_net = net_change / len(entries) if net_change != 0 else 0.0
                                            
                                            assignment_method = "sheet_match"
                                            assignment_reason = f"Assigned from sheet '{sheet_name}' using department matching logic"
                                            
                                            line_item = {
                                                "id": f"{final_dept_id}-LINE-{self.line_item_counter:04d}",
                                                "program_id": current_program['id'] if current_program else None,
                                                "department_id": final_dept_id,
                                                "page_no": page_no,
                                                "line_no": line_no,
                                                "description": entry_text[:500],
                                                "amendment_type": 'CREATE',
                                                "original_amount": float(entry_original),
                                                "final_amount": float(entry_final),
                                                "increase": float(entry_increase),
                                                "decrease": float(entry_decrease),
                                                "net_change": float(entry_net),
                                                "percent_change": float((entry_net / entry_original * 100) if entry_original > 0 else 0.0),
                                                "formulas": row_formulas if row_formulas else None,
                                                "excel_row": row_idx,
                                                "excel_sheet": sheet_name,
                                                "source": {
                                                    "file": "Annex A - Line By Line Amendments.xlsx",
                                                    "sheet": sheet_name,
                                                    "row": row_idx,
                                                    "cell_reference": f"{self.col_num_to_letter(col_mapping.get('line item', 2) + 1)}{row_idx}" if col_mapping.get('line item') is not None else f"A{row_idx}",
                                                    "columns": {
                                                        "description": self.col_num_to_letter(col_mapping.get('line item', 2) + 1) if col_mapping.get('line item') is not None else "C",
                                                        "gab": self.col_num_to_letter(col_mapping.get('gab', 6) + 1) if col_mapping.get('gab') is not None else "G",
                                                        "committee": self.col_num_to_letter(col_mapping.get('committee report', 7) + 1) if col_mapping.get('committee report') is not None else "H",
                                                        "sv_committee": self.col_num_to_letter(col_mapping.get('sv committee', 10) + 1) if col_mapping.get('sv committee') is not None else "K"
                                                    }
                                                },
                                                "assignment": {
                                                    "method": assignment_method,
                                                    "reason": assignment_reason,
                                                    "original_sheet_dept": dept_id,
                                                    "final_dept": final_dept_id
                                                },
                                                "is_split_entry": True,
                                                "split_from": line_item_text[:100]
                                            }
                                            
                                            line_items_data.append(line_item)
                                            if current_program:
                                                program_line_items.append(line_item)
                                                current_program['original_amount'] += entry_original
                                                current_program['final_amount'] += entry_final
                                                current_program['increase'] += entry_increase
                                                current_program['decrease'] += entry_decrease
                                                current_program['net_change'] += entry_net
                                        
                                        continue  # Skip processing the original combined row
                            
                            # If no pipe-separated entries, just mark insertion context for next rows
                            in_insertion_context = True
                            continue  # Skip processing this description row
                        
                        # If we're in insertion context and the line item has pipe-separated entries, split them
                        # Example: "FINANCIAL ASSISTANCE/SUBSIDY | Communication Expenses | Extraordinary and Miscel..."
                        # Should become separate CREATE entries
                        if in_insertion_context and has_gk_values and '|' in line_item_text:
                            # Split by pipe and create separate line items
                            entries = [e.strip() for e in line_item_text.split('|') if e.strip() and len(e.strip()) > 2]
                            
                            if len(entries) > 1:
                                # Create separate line items for each entry
                                for entry_text in entries:
                                    self.line_item_counter += 1
                                    final_dept_id = current_program.get('department_id', dept_id) if current_program else dept_id
                                    
                                    # Split amounts equally among entries
                                    entry_original = original_gab / len(entries) if original_gab > 0 else 0.0
                                    entry_final = final_amount / len(entries) if final_amount > 0 else 0.0
                                    entry_increase = increase / len(entries) if increase > 0 else 0.0
                                    entry_decrease = decrease / len(entries) if decrease < 0 else 0.0
                                    entry_net = net_change / len(entries) if net_change != 0 else 0.0
                                    
                                    assignment_method = "sheet_match"
                                    assignment_reason = f"Assigned from sheet '{sheet_name}' using department matching logic"
                                    
                                    line_item = {
                                        "id": f"{final_dept_id}-LINE-{self.line_item_counter:04d}",
                                        "program_id": current_program['id'] if current_program else None,
                                        "department_id": final_dept_id,
                                        "page_no": page_no,
                                        "line_no": line_no,
                                        "description": entry_text[:500],
                                        "amendment_type": 'CREATE',  # All entries in insertion context are CREATE
                                        "original_amount": float(entry_original),
                                        "final_amount": float(entry_final),
                                        "increase": float(entry_increase),
                                        "decrease": float(entry_decrease),
                                        "net_change": float(entry_net),
                                        "percent_change": float((entry_net / entry_original * 100) if entry_original > 0 else 0.0),
                                        "formulas": row_formulas if row_formulas else None,
                                        "excel_row": row_idx,
                                        "excel_sheet": sheet_name,
                                        "source": {
                                            "file": "Annex A - Line By Line Amendments.xlsx",
                                            "sheet": sheet_name,
                                            "row": row_idx,
                                            "cell_reference": f"{self.col_num_to_letter(col_mapping.get('line item', 2) + 1)}{row_idx}" if col_mapping.get('line item') is not None else f"A{row_idx}",
                                            "columns": {
                                                "description": self.col_num_to_letter(col_mapping.get('line item', 2) + 1) if col_mapping.get('line item') is not None else "C",
                                                "gab": self.col_num_to_letter(col_mapping.get('gab', 6) + 1) if col_mapping.get('gab') is not None else "G",
                                                "committee": self.col_num_to_letter(col_mapping.get('committee report', 7) + 1) if col_mapping.get('committee report') is not None else "H",
                                                "sv_committee": self.col_num_to_letter(col_mapping.get('sv committee', 10) + 1) if col_mapping.get('sv committee') is not None else "K"
                                            }
                                        },
                                        "assignment": {
                                            "method": assignment_method,
                                            "reason": assignment_reason,
                                            "original_sheet_dept": dept_id,
                                            "final_dept": final_dept_id
                                        },
                                        "is_split_entry": True,
                                        "split_from": line_item_text[:100]
                                    }
                                    
                                    line_items_data.append(line_item)
                                    if current_program:
                                        program_line_items.append(line_item)
                                        current_program['original_amount'] += entry_original
                                        current_program['final_amount'] += entry_final
                                        current_program['increase'] += entry_increase
                                        current_program['decrease'] += entry_decrease
                                        current_program['net_change'] += entry_net
                                
                                # Reset insertion context after processing
                                in_insertion_context = False
                                continue  # Skip processing the original combined row
                        
                        if is_grouping_indicator:
                            # Reset activity tracking when a new grouping indicator is found
                            current_activity_name = None
                            current_activity_page = None
                            current_activity_line = None
                            pending_components = []
                            current_header = None
                            continue  # Skip processing grouping indicator rows
                        
                        # Check for strikethrough (indicates deletion)
                        has_strikethrough = False
                        try:
                            for col_idx in range(3, min(7, ws_val.max_column + 1)):
                                cell = ws_val.cell(row_idx, col_idx)
                                if hasattr(cell, 'font') and cell.font and hasattr(cell.font, 'strike'):
                                    if cell.font.strike:
                                        has_strikethrough = True
                                        break
                        except:
                            pass
                        
                        # Use retained values
                        page_no = stored_page_no
                        line_no = stored_line_no
                        category_code = stored_agency_code
                        category_name = stored_agency_name
                        
                        # Columns G-K: Amounts
                        # G = GAB, H = Increase, I = Decrease, J = Net Change, K = Committee
                        col_g_val = ws_val.cell(row_idx, 7).value if ws_val.max_column >= 7 else None  # GAB
                        col_h_val = ws_val.cell(row_idx, 8).value if ws_val.max_column >= 8 else None  # Increase
                        col_i_val = ws_val.cell(row_idx, 9).value if ws_val.max_column >= 9 else None  # Decrease
                        col_j_val = ws_val.cell(row_idx, 10).value if ws_val.max_column >= 10 else None  # Net Change
                        col_k_val = ws_val.cell(row_idx, 11).value if ws_val.max_column >= 11 else None  # Committee
                        
                        # Check if G-K have values
                        has_gk_values = any(
                            val is not None and str(val).strip() not in ['', '-', 'None']
                            for val in [col_g_val, col_h_val, col_i_val, col_j_val, col_k_val]
                        )
                        
                        # Check if this is a header row (no G-K values, has description)
                        is_header_row = not has_gk_values and line_item_text
                        
                        # Check if description contains "TOTAL"
                        is_total_row = 'total' in line_item_text_lower and len(line_item_text.strip()) < 15
                        
                        # Initialize amounts
                        original_gab = 0.0
                        increase = 0.0
                        decrease = 0.0
                        net_change = 0.0
                        final_amount = 0.0
                        committee_amt = 0.0
                        sv_committee = 0.0
                            
                        # Read columns G-K (indices 6-10, 0-based: 7-11)
                        # G = GAB (column 7, index 6)
                        # H = Increase (column 8, index 7) 
                        # I = Decrease (column 9, index 8)
                        # J = Net Change (column 10, index 9)
                        # K = Committee Report (column 11, index 10)
                        
                        # Check if G-K are empty (indicates header row)
                        col_g_val = ws_val.cell(row_idx, 7).value if ws_val.max_column >= 7 else None
                        col_h_val = ws_val.cell(row_idx, 8).value if ws_val.max_column >= 8 else None
                        col_i_val = ws_val.cell(row_idx, 9).value if ws_val.max_column >= 9 else None
                        col_j_val = ws_val.cell(row_idx, 10).value if ws_val.max_column >= 10 else None
                        col_k_val = ws_val.cell(row_idx, 11).value if ws_val.max_column >= 11 else None
                        
                        # Check if this is a header (G-K all empty)
                        is_header_row = (
                            (col_g_val is None or col_g_val == '' or str(col_g_val).strip() == '-' or str(col_g_val).strip() == '') and
                            (col_h_val is None or col_h_val == '' or str(col_h_val).strip() == '-' or str(col_h_val).strip() == '') and
                            (col_i_val is None or col_i_val == '' or str(col_i_val).strip() == '-' or str(col_i_val).strip() == '') and
                            (col_j_val is None or col_j_val == '' or str(col_j_val).strip() == '-' or str(col_j_val).strip() == '') and
                            (col_k_val is None or col_k_val == '' or str(col_k_val).strip() == '-' or str(col_k_val).strip() == '')
                        )
                        
                        # Check if G-K have values (indicates this row has amendment data)
                        has_gk_values = (
                            (col_g_val is not None and col_g_val != '' and str(col_g_val).strip() != '-' and str(col_g_val).strip() != '') or
                            (col_h_val is not None and col_h_val != '' and str(col_h_val).strip() != '-' and str(col_h_val).strip() != '') or
                            (col_i_val is not None and col_i_val != '' and str(col_i_val).strip() != '-' and str(col_i_val).strip() != '') or
                            (col_j_val is not None and col_j_val != '' and str(col_j_val).strip() != '-' and str(col_j_val).strip() != '') or
                            (col_k_val is not None and col_k_val != '' and str(col_k_val).strip() != '-' and str(col_k_val).strip() != '')
                        )
                        
                        if col_mapping['gab'] is not None:
                            gab_val = ws_val.cell(row_idx, col_mapping['gab'] + 1).value
                            original_gab = self.parse_amount(gab_val) * 1000 if gab_val and str(gab_val).strip() not in ['', '-', 'None'] else 0.0
                        elif ws_val.max_column >= 7:
                            gab_val = col_g_val
                            if gab_val and str(gab_val).strip() not in ['', '-', 'None']:
                                if isinstance(gab_val, (int, float)):
                                    original_gab = self.parse_amount(gab_val) * 1000
                                else:
                                    original_gab = 0.0
                            else:
                                original_gab = 0.0
                        else:
                            original_gab = 0.0
                                
                        if col_mapping['committee report'] is not None:
                            comm_val = ws_val.cell(row_idx, col_mapping['committee report'] + 1).value
                            committee_amt = self.parse_amount(comm_val) * 1000 if comm_val and str(comm_val).strip() not in ['', '-', 'None'] else 0.0
                        elif ws_val.max_column >= 8:
                            comm_val = col_h_val  # H is increase, not committee - need to check K
                            if comm_val and str(comm_val).strip() not in ['', '-', 'None']:
                                if isinstance(comm_val, (int, float)):
                                    committee_amt = self.parse_amount(comm_val) * 1000
                                else:
                                    committee_amt = 0.0
                            else:
                                committee_amt = 0.0
                        else:
                            committee_amt = 0.0
                                
                        if col_mapping['sv committee'] is not None:
                            sv_val = ws_val.cell(row_idx, col_mapping['sv committee'] + 1).value
                            sv_committee = self.parse_amount(sv_val) * 1000 if sv_val and str(sv_val).strip() not in ['', '-', 'None'] else 0.0
                        elif ws_val.max_column >= 11:
                            sv_val = col_k_val  # K is Committee Report
                            if sv_val and str(sv_val).strip() not in ['', '-', 'None']:
                                if isinstance(sv_val, (int, float)):
                                    sv_committee = self.parse_amount(sv_val) * 1000
                                else:
                                    sv_committee = 0.0
                            else:
                                sv_committee = 0.0
                        else:
                            sv_committee = 0.0
                        
                        # Also read Increase (H) and Net Change (J) directly
                        increase_val = col_h_val if ws_val.max_column >= 8 else None
                        decrease_val = col_i_val if ws_val.max_column >= 9 else None
                        net_change_val = col_j_val if ws_val.max_column >= 10 else None
                        
                        increase_amt = 0.0
                        if increase_val and str(increase_val).strip() not in ['', '-', 'None']:
                            if isinstance(increase_val, (int, float)):
                                increase_amt = self.parse_amount(increase_val) * 1000
                        
                        decrease_amt = 0.0
                        if decrease_val and str(decrease_val).strip() not in ['', '-', 'None']:
                            if isinstance(decrease_val, (int, float)):
                                decrease_amt = self.parse_amount(decrease_val) * 1000
                        
                        net_change_amt = 0.0
                        if net_change_val and str(net_change_val).strip() not in ['', '-', 'None']:
                            if isinstance(net_change_val, (int, float)):
                                net_change_amt = self.parse_amount(net_change_val) * 1000
                        
                        # Extract FORMULAS
                        row_formulas = {}
                        for col_idx, cell in enumerate(row_form):
                            if cell.data_type == 'f':
                                # Map column index to name if possible
                                col_name = f"col_{col_idx}"
                                for k, v in col_mapping.items():
                                    if v == col_idx:
                                        col_name = k
                                        break
                                row_formulas[col_name] = cell.value
                        
                        # Check if this is a TOTAL row (TOTAL in columns C-F)
                        is_total_row = False
                        # Check columns C-F for "TOTAL" keyword
                        for col_idx in range(3, min(7, ws_val.max_column + 1)):  # Columns C-F (indices 3-6, 1-based)
                            cell_val = ws_val.cell(row_idx, col_idx).value
                            if cell_val and 'total' in str(cell_val).lower().strip():
                                is_total_row = True
                                break
                        
                        # If G-K have values and C-F is NOT TOTAL, this is a component row
                        is_component_row = has_gk_values and not is_total_row and not is_header_row
                        
                        # Determine final amount and changes
                        # For TOTAL rows: if GAB is empty/dash, it's CREATE, use Committee (K) as final
                        if is_total_row and (original_gab == 0 or str(col_g_val).strip() in ['', '-']):
                            # CREATE: GAB empty, final is in Committee (K)
                            final_amount = sv_committee if sv_committee > 0 else 0.0
                            # Use Increase (H) and Net Change (J) from columns directly
                            increase = increase_amt
                            net_change = net_change_amt
                            decrease = abs(decrease_amt) if decrease_amt < 0 else 0.0
                        else:
                            # Normal processing
                            if sv_committee > 0: final_amount = sv_committee
                            elif committee_amt > 0: final_amount = committee_amt
                            elif original_gab > 0: final_amount = original_gab
                            else: final_amount = 0.0
                            
                            increase = max(0, final_amount - original_gab) if final_amount > original_gab else 0
                            decrease = min(0, final_amount - original_gab) if final_amount < original_gab else 0
                            net_change = final_amount - original_gab
                            
                            # Use direct column values if available and more accurate
                            if increase_amt > 0 and abs(increase_amt - increase) > 1000:
                                increase = increase_amt
                            if net_change_amt != 0 and abs(net_change_amt - net_change) > 1000:
                                net_change = net_change_amt
                                final_amount = original_gab + net_change
                        
                        # React based on what values are read
                        # 1. Header row (no G-K values, has description): This is a program/activity name
                        if is_header_row and line_item_text and not is_grouping_indicator:
                            # If we have pending components from previous header, finalize them first
                            if pending_components and current_header:
                                # Finalize previous components as an amendment (no TOTAL found)
                                self._finalize_component_group(
                                    current_header, pending_components, dept_id, current_program,
                                    sheet_name, col_mapping, line_items_data, program_line_items,
                                    page_no, line_no, category_code, category_name
                                )
                            
                            # This is a new program/activity name - store it for grouping components
                            current_header = line_item_text
                            current_activity_name = line_item_text
                            current_activity_page = page_no
                            current_activity_line = line_no
                            # Clear any pending components (new header means new group)
                            pending_components = []
                            
                            # Create a program entry for this activity if it doesn't exist
                            if not current_program or current_program.get('name') != line_item_text:
                                if current_program and program_line_items:
                                    current_program['line_items'] = program_line_items
                                    programs_data.append(current_program)
                                
                                self.prog_counter += 1
                                program_dept_id = dept_id
                                
                                current_program = {
                                    "id": f"{program_dept_id}-PROG-{self.prog_counter:04d}",
                                    "department_id": program_dept_id,
                                    "department_name": next((d['name'] for d in self.departments if d['id'] == program_dept_id), 'UNKNOWN'),
                                    "name": line_item_text,
                                    "grouping_type": current_grouping_type,
                                    "original_amount": 0.0,
                                    "increase": 0.0,
                                    "decrease": 0.0,
                                    "net_change": 0.0,
                                    "final_amount": 0.0,
                                    "percent_change": 0.0,
                                    "project_count": 0,
                                    "page_reference": str(page_no) if page_no else None,
                                    "line_reference": str(line_no) if line_no else None,
                                    "source_sheet": sheet_name,
                                    "line_items": []
                                }
                                program_line_items = []
                            
                            continue  # Skip processing header row itself
                        
                        # Skip empty/note rows
                        line_item_text_lower = line_item_text.lower()
                        is_note_row = (
                            'note:' in line_item_text_lower or
                            line_item_text_lower.startswith('note') or
                            (line_item_text_lower.startswith('subtotal') and not is_total_row) or
                            (not line_item_text and original_gab == 0 and final_amount == 0 and not is_total_row)
                        )
                        
                        # Check for multi-line DELETE pattern: "From lines X to Y, delete the following"
                        is_multiline_delete_start = (
                            'from lines' in line_item_text_lower and 
                            'delete' in line_item_text_lower and 
                            'following' in line_item_text_lower
                        )
                        
                        # Check if this is a TOTAL row (end of multi-line DELETE) - already set above, but check again
                        if not is_total_row:
                            is_total_row = (
                                line_item_text_lower.strip() == 'total' or
                                (line_item_text_lower.startswith('total') and len(line_item_text.strip()) < 10)
                            )
                        
                        # Handle multi-line DELETE grouping
                        if is_multiline_delete_start:
                            # Start a new multi-line DELETE group
                            pending_multiline_delete = {
                                'page_no': page_no,
                                'line_no': line_no,
                                'description_parts': [line_item_text],
                                'original_amount': original_gab,
                                'final_amount': final_amount,
                                'increase': increase,
                                'decrease': decrease,
                                'net_change': net_change,
                                'row_start': row_idx,
                                'row_formulas': row_formulas
                            }
                            continue  # Skip this row, we'll process it when we hit TOTAL
                        
                        elif pending_multiline_delete:
                            # We're in a multi-line DELETE group
                            if is_total_row:
                                # End of multi-line DELETE - create the combined line item
                                combined_description = ' | '.join(pending_multiline_delete['description_parts'])
                                
                                # Create the DELETE line item
                                self.line_item_counter += 1
                                final_dept_id = current_program.get('department_id', dept_id) if current_program else dept_id
                                
                                # Assignment logic
                                assignment_method = "sheet_match"
                                assignment_reason = f"Assigned from sheet '{sheet_name}' using department matching logic"
                                
                                line_item = {
                                    "id": f"{final_dept_id}-LINE-{self.line_item_counter:04d}",
                                    "program_id": current_program['id'] if current_program else None,
                                    "department_id": final_dept_id,
                                    "page_no": page_no,  # Use current stored page_no
                                    "line_no": line_no,  # Use current stored line_no
                                    "description": combined_description[:500],
                                    "amendment_type": 'DELETE',
                                    "original_amount": float(pending_multiline_delete['original_amount']),
                                    "final_amount": float(pending_multiline_delete['final_amount']),
                                    "increase": float(pending_multiline_delete['increase']),
                                    "decrease": float(pending_multiline_delete['decrease']),
                                    "net_change": float(pending_multiline_delete['net_change']),
                                    "percent_change": float((pending_multiline_delete['net_change'] / pending_multiline_delete['original_amount'] * 100) if pending_multiline_delete['original_amount'] > 0 else 0.0),
                                    "formulas": pending_multiline_delete['row_formulas'] if pending_multiline_delete['row_formulas'] else None,
                                    "excel_row": pending_multiline_delete['row_start'],
                                    "excel_sheet": sheet_name,
                                    "source": {
                                        "file": "Annex A - Line By Line Amendments.xlsx",
                                        "sheet": sheet_name,
                                        "row": pending_multiline_delete['row_start'],
                                        "cell_reference": f"{self.col_num_to_letter(col_mapping.get('line item', 2) + 1)}{pending_multiline_delete['row_start']}" if col_mapping.get('line item') is not None else f"A{pending_multiline_delete['row_start']}",
                                        "columns": {
                                            "description": self.col_num_to_letter(col_mapping.get('line item', 2) + 1) if col_mapping.get('line item') is not None else "C",
                                            "gab": self.col_num_to_letter(col_mapping.get('gab', 6) + 1) if col_mapping.get('gab') is not None else "G",
                                            "committee": self.col_num_to_letter(col_mapping.get('committee report', 7) + 1) if col_mapping.get('committee report') is not None else "H",
                                            "sv_committee": self.col_num_to_letter(col_mapping.get('sv committee', 10) + 1) if col_mapping.get('sv committee') is not None else "K"
                                        }
                                    },
                                    "assignment": {
                                        "method": assignment_method,
                                        "reason": assignment_reason,
                                        "original_sheet_dept": dept_id,
                                        "final_dept": final_dept_id
                                    },
                                    "is_multiline": True,
                                    "multiline_count": len(pending_multiline_delete['description_parts'])
                                }
                                
                                line_items_data.append(line_item)
                                if current_program:
                                    program_line_items.append(line_item)
                                    current_program['original_amount'] += pending_multiline_delete['original_amount']
                                    current_program['final_amount'] += pending_multiline_delete['final_amount']
                                    current_program['increase'] += pending_multiline_delete['increase']
                                    current_program['decrease'] += pending_multiline_delete['decrease']
                                
                                pending_multiline_delete = None
                                # Reset stored page_no/line_no after finalizing multi-line DELETE
                                stored_page_no = None
                                stored_line_no = None
                                continue  # Skip the TOTAL row itself
                            else:
                                # Add this row to the multi-line DELETE group
                                if line_item_text and not is_note_row:
                                    pending_multiline_delete['description_parts'].append(line_item_text)
                                    # Accumulate amounts (use max for original, sum for others)
                                    pending_multiline_delete['original_amount'] = max(pending_multiline_delete['original_amount'], original_gab)
                                    pending_multiline_delete['final_amount'] = max(pending_multiline_delete['final_amount'], final_amount)
                                    pending_multiline_delete['increase'] = max(pending_multiline_delete['increase'], increase)
                                    pending_multiline_delete['decrease'] = min(pending_multiline_delete['decrease'], decrease)
                                    pending_multiline_delete['net_change'] = pending_multiline_delete['final_amount'] - pending_multiline_delete['original_amount']
                                continue  # Skip processing this row individually
                        
                        if is_note_row: continue
                        
                        # Handle component grouping logic (for program/activity structure)
                        # 2. Component row (has G-K values, not TOTAL, and a current_header exists)
                        if has_gk_values and not is_total_row and current_header:
                            # This is a component of the current activity/program
                            pending_components.append({
                                'description': line_item_text,
                                'original_gab': original_gab,
                                'final_amount': final_amount,
                                'increase': increase,
                                'decrease': decrease,
                                'net_change': net_change,
                                'row_idx': row_idx,
                                'row_formulas': row_formulas,
                                'page_no': page_no,
                                'line_no': line_no,
                                'agency_code': category_code,
                                'agency_name': category_name
                            })
                            continue  # Skip processing this row individually, it's a component
                        
                        # 3. TOTAL row (ends a component group)
                        if is_total_row:
                            if pending_components and current_header:
                                # Finalize the component group with TOTAL
                                self._finalize_component_group(
                                    current_header, pending_components, dept_id, current_program,
                                    sheet_name, col_mapping, line_items_data, program_line_items,
                                    page_no, line_no, category_code, category_name
                                )
                                pending_components = []
                                current_header = None
                                current_activity_name = None
                                current_activity_page = None
                                current_activity_line = None
                                continue  # Skip the TOTAL row itself - it's just an indicator
                            # If no pending components, TOTAL row is a standalone line item
                            # (fall through to normal line item processing)
                        
                        # 4. Standalone line item (has G-K values, no current_header, not TOTAL)
                        # OR: TOTAL row with no pending components
                        # These will be processed as normal line items below
                        
                        # Determine amendment type
                        # Be conservative: only classify as amendment if there's clear evidence of change
                        amendment_type = 'RETAIN'  # Default: retained (not an amendment)
                        
                        # Check for actual amendments (items with clear changes)
                        # TOTAL rows with empty GAB are CREATE
                        if is_total_row and (original_gab == 0 or str(col_g_val).strip() in ['', '-']):
                            amendment_type = 'CREATE'
                        # Items that follow "Between lines X and Y, insert..." or "After line X, insert..." are CREATE
                        # (The description row itself is already skipped above)
                        elif has_gk_values and (original_gab == 0 or str(col_g_val).strip() in ['', '-']) and (final_amount > 0 or increase > 0):
                            # New items (no original amount, but has final/increase) are CREATE
                            amendment_type = 'CREATE'
                        elif any(k in line_item_text_lower for k in ['delete', 'remove', 'in lieu', 'strike out']) or has_strikethrough:
                            amendment_type = 'DELETE'
                        elif net_change > 0:
                            amendment_type = 'INCREASE'
                        elif net_change < 0:
                            amendment_type = 'DECREASE'
                        elif original_gab == final_amount and original_gab > 0:
                            # Same non-zero amount - check if there's clear evidence of rename
                            # Look for explicit rename/reclassify language
                            rename_keywords = ['rename', 'reclass', 'reclassify', 'reclassification', 'change to', 'changed to', 
                                             'formerly', 'previously', 'now', 'from', 'to be', 'shall be']
                            if any(k in line_item_text_lower for k in rename_keywords):
                                amendment_type = 'RENAME'
                            # Otherwise, it's RETAIN (no clear evidence of change)
                        # If original_gab == 0 and final_amount == 0, it's already RETAIN (default)
                        # BUT: If there's a page/line reference, it might still be an amendment (organizational reference)
                        # Only classify as RETAIN if it's truly just organizational with no changes
                        if amendment_type == 'RETAIN' and (page_no or line_no) and len(line_item_text) > 10:
                            # Has page/line reference and description - might be organizational, but check if it has any amount
                            if original_gab > 0 or final_amount > 0:
                                # Has amounts but no change - could be RENAME if description suggests it
                                if any(k in line_item_text_lower for k in ['change', 'modify', 'update', 'revise']):
                                    amendment_type = 'RENAME'
                        
                        # Check for program header
                        is_program_header = (
                            len(line_item_text) > 15 and 
                            (original_gab == 0 or original_gab < 1000000) and
                            (line_item_text.startswith(('I.', 'II.', 'III.', 'IV.', 'V.', 'VI.', 'VII.', 'VIII.', 'IX.', 'X.')) or
                             ('program' in line_item_text_lower and 'by activity' not in line_item_text_lower and 'by project' not in line_item_text_lower) or 
                             'sub-program' in line_item_text_lower or
                             ('activity' in line_item_text_lower and 'by' not in line_item_text_lower and len(line_item_text) > 20))
                        )
                        
                        # Check for line item
                        # TOTAL rows are always line items (amendments)
                        is_line_item = is_total_row or (
                            (original_gab > 0 or final_amount > 0 or increase > 0 or decrease < 0) and
                            len(line_item_text) > 3 and
                            not is_program_header and
                            not line_item_text_lower.startswith(('note:', 'subtotal'))
                        ) or (
                            (page_no or line_no) and len(line_item_text) > 3 and not is_program_header
                        )
                        
                        if is_program_header:
                            if current_program and program_line_items:
                                current_program['line_items'] = program_line_items
                                programs_data.append(current_program)
                            
                            self.prog_counter += 1
                            program_dept_id = dept_id
                            
                            # OEOs special handling
                            if sheet_name.upper() == 'OEOs' or 'EXECUTIVE' in sheet_name.upper():
                                if any(kw in line_item_text_lower for kw in ['movie', 'television', 'review', 'classification']):
                                    oma_dept = next((d for d in self.departments if d.get('code', '').upper() == 'OMA'), None)
                                    if oma_dept: program_dept_id = oma_dept['id']
                            
                            dept_name = next((d['name'] for d in self.departments if d['id'] == program_dept_id), 'UNKNOWN')
                            
                            current_program = {
                                "id": f"{program_dept_id}-PROG-{self.prog_counter:04d}",
                                "department_id": program_dept_id,
                                "department_name": dept_name,
                                "name": line_item_text,
                                "original_amount": 0.0,
                                "increase": 0.0,
                                "decrease": 0.0,
                                "net_change": 0.0,
                                "final_amount": 0.0,
                                "percent_change": 0.0,
                                "project_count": 0,
                                "page_reference": str(page_no) if page_no else None,
                                "source_sheet": sheet_name,
                                "line_items": []
                            }
                            program_line_items = []
                        
                        elif is_line_item:
                            # Clean description: remove "Between lines X and Y, insert..." prefix if present
                            cleaned_description = line_item_text
                            # Remove insertion description prefix
                            cleaned_description = re.sub(r'^.*?between\s+lines?\s+\d+\s+and\s+\d+.*?insert\s+(the\s+following\s+)?(objects?\s+of\s+expenditure|items?|activities?|projects?):?\s*', '', cleaned_description, flags=re.IGNORECASE)
                            cleaned_description = re.sub(r'^.*?after\s+line\s+\d+.*?insert\s+(the\s+following\s+)?(objects?\s+of\s+expenditure|items?|activities?|projects?):?\s*', '', cleaned_description, flags=re.IGNORECASE)
                            cleaned_description = cleaned_description.strip()
                            
                            # Check if cleaned description has pipe-separated entries (even if not in insertion context)
                            # Example: "FINANCIAL ASSISTANCE/SUBSIDY | Communication Expenses | Extraordinary and Miscel..."
                            if '|' in cleaned_description and has_gk_values:
                                # Split into separate entries
                                entries = [e.strip() for e in cleaned_description.split('|') if e.strip() and len(e.strip()) > 2]
                                
                                if len(entries) > 1:
                                    # Create separate line items for each entry
                                    for entry_text in entries:
                                        self.line_item_counter += 1
                                        final_dept_id = current_program.get('department_id', dept_id) if current_program else dept_id
                                        
                                        # Use stored page_no and line_no
                                        amendment_page_no = page_no
                                        amendment_line_no = line_no
                                        
                                        # Split amounts equally among entries
                                        entry_original = original_gab / len(entries) if original_gab > 0 else 0.0
                                        entry_final = final_amount / len(entries) if final_amount > 0 else 0.0
                                        entry_increase = increase / len(entries) if increase > 0 else 0.0
                                        entry_decrease = decrease / len(entries) if decrease < 0 else 0.0
                                        entry_net = net_change / len(entries) if net_change != 0 else 0.0
                                        
                                        # Determine amendment type for split entries
                                        entry_amendment_type = amendment_type
                                        if entry_original == 0 and (entry_final > 0 or entry_increase > 0):
                                            entry_amendment_type = 'CREATE'
                                        
                                        assignment_method = "sheet_match"
                                        assignment_reason = f"Assigned from sheet '{sheet_name}' using department matching logic"
                                        
                                        line_item = {
                                            "id": f"{final_dept_id}-LINE-{self.line_item_counter:04d}",
                                            "program_id": current_program['id'] if current_program else None,
                                            "department_id": final_dept_id,
                                            "page_no": amendment_page_no,
                                            "line_no": amendment_line_no,
                                            "description": entry_text[:500],
                                            "amendment_type": entry_amendment_type,
                                            "original_amount": float(entry_original),
                                            "final_amount": float(entry_final),
                                            "increase": float(entry_increase),
                                            "decrease": float(entry_decrease),
                                            "net_change": float(entry_net),
                                            "percent_change": float((entry_net / entry_original * 100) if entry_original > 0 else 0.0),
                                            "formulas": row_formulas if row_formulas else None,
                                            "excel_row": row_idx,
                                            "excel_sheet": sheet_name,
                                            "source": {
                                                "file": "Annex A - Line By Line Amendments.xlsx",
                                                "sheet": sheet_name,
                                                "row": row_idx,
                                                "cell_reference": f"{self.col_num_to_letter(col_mapping.get('line item', 2) + 1)}{row_idx}" if col_mapping.get('line item') is not None else f"A{row_idx}",
                                                "columns": {
                                                    "description": self.col_num_to_letter(col_mapping.get('line item', 2) + 1) if col_mapping.get('line item') is not None else "C",
                                                    "gab": self.col_num_to_letter(col_mapping.get('gab', 6) + 1) if col_mapping.get('gab') is not None else "G",
                                                    "committee": self.col_num_to_letter(col_mapping.get('committee report', 7) + 1) if col_mapping.get('committee report') is not None else "H",
                                                    "sv_committee": self.col_num_to_letter(col_mapping.get('sv committee', 10) + 1) if col_mapping.get('sv committee') is not None else "K"
                                                }
                                            },
                                            "assignment": {
                                                "method": assignment_method,
                                                "reason": assignment_reason,
                                                "original_sheet_dept": dept_id,
                                                "final_dept": final_dept_id
                                            },
                                            "is_split_entry": True,
                                            "split_from": line_item_text[:100]
                                        }
                                        
                                        line_items_data.append(line_item)
                                        if current_program:
                                            program_line_items.append(line_item)
                                            current_program['original_amount'] += entry_original
                                            current_program['final_amount'] += entry_final
                                            current_program['increase'] += entry_increase
                                            current_program['decrease'] += entry_decrease
                                            current_program['net_change'] += entry_net
                                    
                                    # Reset insertion context after processing
                                    in_insertion_context = False
                                    continue  # Skip processing the original combined row
                            
                            # Regular single line item (not split)
                            self.line_item_counter += 1
                            final_dept_id = current_program.get('department_id', dept_id) if current_program else dept_id
                            
                            # Use stored page_no and line_no for this amendment
                            amendment_page_no = page_no  # Use current stored page_no
                            amendment_line_no = line_no  # Use current stored line_no
                            
                            # If this is a TOTAL row, reset stored values for next amendment
                            if is_total_row:
                                stored_page_no = None
                                stored_line_no = None
                            
                            # Reset insertion context after processing a line item
                            in_insertion_context = False
                            
                            # Determine how this line item was assigned to this department
                            assignment_method = "direct"
                            assignment_reason = None
                            if sheet_name.upper() == 'OEOs' or 'EXECUTIVE' in sheet_name.upper():
                                if current_program and any(kw in current_program.get('name', '').lower() for kw in ['movie', 'television', 'review', 'classification']):
                                    assignment_method = "program_name_match"
                                    assignment_reason = f"Matched via program name '{current_program.get('name', '')[:50]}' containing movie/television keywords"
                            elif final_dept_id != dept_id:
                                assignment_method = "reassigned"
                                assignment_reason = f"Reassigned from sheet default department to {final_dept_id} based on content analysis"
                            else:
                                assignment_method = "sheet_match"
                                assignment_reason = f"Assigned from sheet '{sheet_name}' using department matching logic"
                            
                            # Determine source file name
                            source_file = "Annex A - Line By Line Amendments.xlsx"
                            if sheet_name in ['DA', 'DepED', 'DOTR']:
                                # These might come from other files, but for now all Annex A
                                pass
                            
                            line_item = {
                                "id": f"{final_dept_id}-LINE-{self.line_item_counter:04d}",
                                "program_id": current_program['id'] if current_program else None,
                                "department_id": final_dept_id,
                                "page_no": amendment_page_no,
                                "line_no": amendment_line_no,
                                "description": cleaned_description[:500],  # Use cleaned description
                                "amendment_type": amendment_type,
                                "original_amount": float(original_gab),
                                "final_amount": float(final_amount),
                                "increase": float(increase),
                                "decrease": float(decrease),
                                "net_change": float(net_change),
                                "percent_change": float((net_change / original_gab * 100) if original_gab > 0 else 0.0),
                                "formulas": row_formulas if row_formulas else None,
                                "excel_row": row_idx,
                                "excel_sheet": sheet_name,
                                # Enhanced traceability fields
                                "source": {
                                    "file": source_file,
                                    "sheet": sheet_name,
                                    "row": row_idx,
                                    "cell_reference": f"{self.col_num_to_letter(col_mapping.get('line item', 2) + 1)}{row_idx}" if col_mapping.get('line item') is not None else f"A{row_idx}",
                                    "columns": {
                                        "description": self.col_num_to_letter(col_mapping.get('line item', 2) + 1) if col_mapping.get('line item') is not None else "C",
                                        "gab": self.col_num_to_letter(col_mapping.get('gab', 6) + 1) if col_mapping.get('gab') is not None else "G",
                                        "committee": self.col_num_to_letter(col_mapping.get('committee report', 7) + 1) if col_mapping.get('committee report') is not None else "H",
                                        "sv_committee": self.col_num_to_letter(col_mapping.get('sv committee', 10) + 1) if col_mapping.get('sv committee') is not None else "K"
                                    }
                                },
                                "assignment": {
                                    "method": assignment_method,
                                    "reason": assignment_reason,
                                    "original_sheet_dept": dept_id,
                                    "final_dept": final_dept_id
                                }
                            }
                            
                            line_items_data.append(line_item)
                            
                            if current_program:
                                program_line_items.append(line_item)
                                current_program['original_amount'] += original_gab
                                current_program['final_amount'] += final_amount
                                current_program['increase'] += increase
                                current_program['decrease'] += decrease
                                current_program['net_change'] += net_change
                    
                    # Finalize any pending components at the end of the sheet
                    if pending_components and current_header:
                        self._finalize_component_group(
                            current_header, pending_components, dept_id, current_program,
                            sheet_name, col_mapping, line_items_data, program_line_items,
                            page_no, line_no, category_code, category_name
                        )
                        pending_components = []
                        current_header = None
                    
                    if current_program and program_line_items:
                        current_program['line_items'] = program_line_items
                        if current_program['original_amount'] > 0:
                            current_program['percent_change'] = (current_program['net_change'] / current_program['original_amount'] * 100)
                        programs_data.append(current_program)
                    
                except Exception as e:
                    print(f"      ⚠️  Error in sheet {sheet_name}: {e}")
                    import traceback
                    traceback.print_exc()
                    continue
            
            wb_values.close()
            wb_formulas.close()
            
            # Add line items to instance variable
            self.line_items.extend(line_items_data)
            
            # Add Annex A departments to main departments list
            # Merge with existing departments from General Summary (prefer Annex A names)
            # Use fuzzy matching to handle slight name variations (e.g., comma differences)
            # IMPORTANT: Prevent duplicates - if a department exists, merge instead of adding duplicate
            for annex_dept in annex_a_departments:
                annex_name_upper = annex_dept.get('name', '').upper().strip()
                annex_id = annex_dept.get('id', '')
                
                # Check for exact match first (by ID, then by name)
                existing = next((d for d in self.departments if d.get('id') == annex_id), None)
                
                if not existing:
                    # Try exact name match
                    existing = next((d for d in self.departments if d.get('name', '').upper().strip() == annex_name_upper), None)
                
                # If no exact match, try fuzzy match (normalize by removing extra commas/spaces)
                if not existing:
                    annex_normalized = re.sub(r'[, ]+', ' ', annex_name_upper)
                    for d in self.departments:
                        dept_normalized = re.sub(r'[, ]+', ' ', d.get('name', '').upper().strip())
                        if annex_normalized == dept_normalized:
                            existing = d
                            break
                
                # If still no match, try matching by ID prefix (e.g., "DEP" matches "DEP")
                if not existing and annex_id:
                    existing = next((d for d in self.departments if d.get('id') == annex_id), None)
                
                if not existing:
                    # Only add if it doesn't already exist
                    # But first, calculate amounts from line items for this Annex A department
                    dept_id = annex_dept.get('id')
                    dept_line_items = [li for li in line_items_data if li.get('department_id') == dept_id]
                    if dept_line_items:
                        annex_dept['original_amount'] = sum(li.get('original_amount', 0) for li in dept_line_items)
                        annex_dept['final_amount'] = sum(li.get('final_amount', 0) for li in dept_line_items)
                        annex_dept['increase'] = sum(li.get('increase', 0) for li in dept_line_items)
                        annex_dept['decrease'] = sum(li.get('decrease', 0) for li in dept_line_items)
                        annex_dept['net_change'] = annex_dept['final_amount'] - annex_dept['original_amount']
                        annex_dept['percent_change'] = (annex_dept['net_change'] / annex_dept['original_amount'] * 100) if annex_dept['original_amount'] > 0 else 0.0
                    self.departments.append(annex_dept)
                else:
                    # Update existing department with Annex A sheet info (merge, don't duplicate)
                    if not existing.get('sheet_name') or existing.get('sheet_name') == 'N/A':
                        existing['sheet_name'] = annex_dept.get('sheet_name')
                    existing['source'] = 'Both'
                    # Prefer Annex A name (it's cleaner, from row 10)
                    if annex_dept.get('name'):
                        existing['name'] = annex_dept.get('name')
                    # IMPORTANT: Keep General Summary amounts if they exist (they represent full budget)
                    # Only calculate from line items if existing department has 0 amounts
                    existing_original = existing.get('original_amount', 0)
                    existing_final = existing.get('final_amount', 0)
                    if existing_original == 0 and existing_final == 0:
                        # Calculate amounts from Annex A line items
                        dept_id = existing.get('id')
                        dept_line_items = [li for li in line_items_data if li.get('department_id') == dept_id]
                        if dept_line_items:
                            annex_original = sum(li.get('original_amount', 0) for li in dept_line_items)
                            annex_final = sum(li.get('final_amount', 0) for li in dept_line_items)
                            annex_increase = sum(li.get('increase', 0) for li in dept_line_items)
                            annex_decrease = sum(li.get('decrease', 0) for li in dept_line_items)
                            existing['original_amount'] = annex_original
                            existing['final_amount'] = annex_final
                            existing['increase'] = annex_increase
                            existing['decrease'] = annex_decrease
                            existing['net_change'] = annex_final - annex_original
                            existing['percent_change'] = (existing['net_change'] / annex_original * 100) if annex_original > 0 else 0.0
                    # If existing has amounts, keep them (General Summary is authoritative for full budget)
                    # Don't overwrite with Annex A amounts as those only represent amended items
            
            print(f"✅ Parsed {len(programs_data)} programs and {len(line_items_data)} line items from Annex A")
            print(f"✅ Created {len(annex_a_departments)} department-level entries from Annex A worksheets")
            
            return programs_data
            
        except Exception as e:
            print(f"❌ Error parsing Annex A: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def parse_da_farm_roads(self) -> List[Dict]:
        """Parse Annex A-1 DA Farm-to-Market Roads"""
        file_path = self.base_dir / "Annex A-1 DA-Farm-to-Market Roads.xlsx"
        
        print(f"🚜 Parsing DA Farm-to-Market Roads...")
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[wb.sheetnames[0]]
            
            projects_data = []
            
            # Find header row (look for "NAME OF PROJECTS" or "AMOUNT")
            header_row = None
            for row_idx in range(1, min(20, ws.max_row + 1)):
                row_values = [str(ws.cell(row_idx, col).value or '').strip() for col in range(1, min(15, ws.max_column + 1))]
                row_str = ' '.join(row_values).lower()
                if ('project' in row_str and 'amount' in row_str) or ('name' in row_str and 'project' in row_str):
                    header_row = row_idx
                    break
                
            if header_row is None:
                print(f"      ⚠️  Could not find header row")
                wb.close()
                return []
            
            # Find column indices
            col_project = None
            col_amount = None
            col_region = None
            
            for col_idx in range(1, ws.max_column + 1):
                cell_val = str(ws.cell(header_row, col_idx).value or '').lower()
                if 'project' in cell_val or 'name' in cell_val:
                    col_project = col_idx
                if 'amount' in cell_val or 'pesos' in cell_val:
                    col_amount = col_idx
                if 'region' in cell_val:
                    col_region = col_idx
            
            # Parse data rows
            current_category = None
            for row_idx in range(header_row + 1, ws.max_row + 1):
                project_name = None
                region = None
                amount = 0.0
                
                # Get project name
                if col_project:
                    project_val = ws.cell(row_idx, col_project).value
                    if project_val:
                        project_name = str(project_val).strip()
                
                # Get amount
                if col_amount:
                    amount_val = ws.cell(row_idx, col_amount).value
                    amount = self.parse_amount(amount_val)
                
                # Get region (might be in project name or separate column)
                if col_region:
                    region_val = ws.cell(row_idx, col_region).value
                    if region_val:
                        region = str(region_val).strip()
                
                # Check if this is a category header (like "REPAIR/ REHABILITATION")
                if project_name and len(project_name) > 10 and amount > 1000000000:  # Large amounts might be category totals
                    if any(keyword in project_name.upper() for keyword in ['REPAIR', 'REHABILITATION', 'CONSTRUCTION', 'REGION']):
                        current_category = project_name
                        continue
                
                # Check if this is a region header
                if project_name and 'Region' in project_name and amount > 100000000:
                    region = project_name
                    continue
                
                # Only create project if we have name and amount
                if project_name and len(project_name) > 3 and amount > 0:
                    # Skip if it looks like a header or total
                    if any(keyword in project_name.upper() for keyword in ['TOTAL', 'SUBTOTAL', 'GRAND', 'STATUS', 'VALIDATION']):
                        continue
                    
                    self.proj_counter += 1
                    
                    # Extract region from name if not in separate column
                    if not region and 'Region' in project_name:
                        parts = project_name.split('-', 1)
                        if len(parts) > 1:
                            region = parts[0].strip()
                            project_name = parts[1].strip()
                    
                    projects_data.append({
                        "id": f"DA-FMR-{self.proj_counter:04d}",
                        "program_id": None,
                        "department_id": "DA",
                        "name": project_name,
                        "description": f"Farm-to-Market Road{(' - ' + current_category) if current_category else ''}",
                        "location": {
                            "region": region,
                            "province": None,
                            "municipality": None,
                            "barangay": None
                        },
                        "original_amount": amount,
                        "final_amount": amount,
                        "net_change": 0.0,
                        "percent_change": 0.0,
                        "amendments": [],
                        "revised_name": None,
                        "page_reference": None,
                        "source_sheet": "Annex A-1"
                    })
            
            wb.close()
            print(f"✅ Parsed {len(projects_data)} farm road projects")
            return projects_data
            
        except Exception as e:
            print(f"❌ Error parsing DA Farm Roads: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def parse_deped_schools(self) -> List[Dict]:
        """Parse Annex A-2 DepEd Schools"""
        file_path = self.base_dir / "Annex A-2 DepEd - Office of the Secretary - Non Implementing Unit Secondary Schools.xlsx"
        
        print(f"🏫 Parsing DepEd Schools...")
        
        try:
            projects_data = []
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Parse Sheet2 which has the budget allocation changes
            if 'Sheet2' in wb.sheetnames:
                ws = wb['Sheet2']
                
                # Find header row (usually row 4 or 5)
                header_row = None
                for row_idx in range(1, min(10, ws.max_row + 1)):
                    row_values = [str(ws.cell(row_idx, col).value or '').strip() for col in range(1, min(15, ws.max_column + 1))]
                    if any('school' in str(v).lower() for v in row_values) or any('amount' in str(v).lower() for v in row_values):
                        header_row = row_idx
                        break
                
                if header_row:
                    # Find column indices
                    col_school = None
                    col_original = None
                    col_final = None
                    
                    for col_idx in range(1, ws.max_column + 1):
                        cell_val = str(ws.cell(header_row, col_idx).value or '').lower()
                        if 'school' in cell_val and col_school is None:
                            col_school = col_idx
                        if 'original' in cell_val or 'gab' in cell_val:
                            col_original = col_idx
                        if 'final' in cell_val or 'committee' in cell_val or 'amendment' in cell_val:
                            col_final = col_idx
                    
                    # Parse data rows
                    for row_idx in range(header_row + 1, ws.max_row + 1):
                        school_name = ws.cell(row_idx, col_school).value if col_school else None
                        if not school_name or not str(school_name).strip():
                            continue
                        
                        original_amount = self.parse_amount(ws.cell(row_idx, col_original).value) if col_original else 0.0
                        final_amount = self.parse_amount(ws.cell(row_idx, col_final).value) if col_final else original_amount
                        
                        if original_amount > 0 or final_amount > 0:
                            self.proj_counter += 1
                            net_change = final_amount - original_amount
                            
                            projects_data.append({
                                "id": f"DepEd-School-{self.proj_counter:04d}",
                                "program_id": None,
                                "department_id": "DepEd",
                                "name": str(school_name).strip(),
                                "description": "Non-Implementing Unit Secondary School",
                                "location": {"region": None, "province": None, "municipality": None, "barangay": None},
                                "original_amount": original_amount,
                                "final_amount": final_amount,
                                "net_change": net_change,
                                "percent_change": (net_change / original_amount * 100) if original_amount > 0 else 0.0,
                                "amendments": [],
                                "revised_name": None,
                                "page_reference": None,
                                "source_sheet": "Sheet2"
                            })
            
            wb.close()
            print(f"✅ Parsed {len(projects_data)} DepEd schools")
            return projects_data
            
        except Exception as e:
            print(f"❌ Error parsing DepEd Schools: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def parse_nia_operations(self) -> Dict:
        """Parse Annex A-4 NIA Operations"""
        file_path = self.base_dir / "Annex A-4 BSGC-OEOs-NIA Details of NIA's Operations Budget.xlsx"
        
        print(f"💧 Parsing NIA Operations...")
        
        try:
            programs_data = []
            projects_data = []
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            if 'Adjusted per ComRep ' in wb.sheetnames:
                ws = wb['Adjusted per ComRep ']
                
                # Find header row
                header_row = None
                for row_idx in range(1, min(10, ws.max_row + 1)):
                    row_values = [str(ws.cell(row_idx, col).value or '').strip() for col in range(1, min(10, ws.max_column + 1))]
                    if any('program' in str(v).lower() for v in row_values) or any('amount' in str(v).lower() for v in row_values):
                        header_row = row_idx
                        break
                
                if header_row:
                    # Parse data rows
                    current_program = None
                    for row_idx in range(header_row + 1, min(ws.max_row + 1, header_row + 500)):  # Limit for performance
                        col_b = ws.cell(row_idx, 2).value
                        col_c = ws.cell(row_idx, 3).value
                        
                        # Check if this is a program header
                        if col_b and isinstance(col_b, str) and ('program' in str(col_b).lower() or 'irrigation' in str(col_b).lower()):
                            if current_program:
                                programs_data.append(current_program)
                            
                            self.prog_counter += 1
                            current_program = {
                                "id": f"NIA-Prog-{self.prog_counter:04d}",
                                "department_id": "NIA",
                                "name": str(col_b).strip(),
                                "description": None,
                                "original_amount": 0.0,
                                "final_amount": 0.0,
                                "increase": 0.0,
                                "decrease": 0.0,
                                "net_change": 0.0,
                                "percent_change": 0.0,
                                "line_items": []
                            }
                        elif current_program and col_c:
                            # This is a project/activity under the program
                            project_name = str(col_c).strip()
                            if len(project_name) > 5:
                                # Try to extract amounts from columns
                                amounts = []
                                for col_idx in range(5, min(10, ws.max_column + 1)):
                                    val = ws.cell(row_idx, col_idx).value
                                    if isinstance(val, (int, float)) and val > 0:
                                        amounts.append(float(val) * 1000)  # Convert from thousands
                                
                                if amounts:
                                    original = amounts[0] if len(amounts) > 0 else 0.0
                                    final = amounts[-1] if len(amounts) > 1 else original
                                    
                                    self.proj_counter += 1
                                    net_change = final - original
                                    
                                    project = {
                                        "id": f"NIA-Proj-{self.proj_counter:04d}",
                                        "program_id": current_program['id'],
                                        "department_id": "NIA",
                                        "name": project_name,
                                        "description": None,
                                        "location": {"region": None, "province": None, "municipality": None, "barangay": None},
                                        "original_amount": original,
                                        "final_amount": final,
                                        "net_change": net_change,
                                        "percent_change": (net_change / original * 100) if original > 0 else 0.0,
                                        "amendments": [],
                                        "revised_name": None,
                                        "page_reference": None,
                                        "source_sheet": "Adjusted per ComRep "
                                    }
                                    
                                    projects_data.append(project)
                                    current_program['original_amount'] += original
                                    current_program['final_amount'] += final
                                    current_program['net_change'] += net_change
                    
                    if current_program:
                        programs_data.append(current_program)
            
            wb.close()
            print(f"✅ Parsed {len(programs_data)} NIA programs and {len(projects_data)} projects")
            return {"programs": programs_data, "projects": projects_data}
            
        except Exception as e:
            print(f"❌ Error parsing NIA Operations: {e}")
            import traceback
            traceback.print_exc()
            return {"programs": [], "projects": []}
    
    def parse_dpwh_projects(self) -> List[Dict]:
        """Parse Annex A-5 DPWH Projects"""
        file_path = self.base_dir / "Annex A-5 Details of DPWH's Programs&Projects.xlsx"
        
        print(f"🛣️  Parsing DPWH Projects...")
        
        try:
            projects_data = []
            wb = openpyxl.load_workbook(file_path, data_only=True)
            
            # Parse the main sheet "DPWH 1-C ComRep"
            if 'DPWH 1-C ComRep' in wb.sheetnames:
                ws = wb['DPWH 1-C ComRep']
                
                # Find header row
                header_row = None
                for row_idx in range(1, min(10, ws.max_row + 1)):
                    row_values = [str(ws.cell(row_idx, col).value or '').strip() for col in range(1, min(21, ws.max_column + 1))]
                    if any('project' in str(v).lower() for v in row_values) or any('activity' in str(v).lower() for v in row_values):
                        header_row = row_idx
                        break
                
                if header_row:
                    # Find column indices
                    col_project = None
                    col_amount = None
                    
                    for col_idx in range(1, ws.max_column + 1):
                        cell_val = str(ws.cell(header_row, col_idx).value or '').lower()
                        if ('project' in cell_val or 'activity' in cell_val) and col_project is None:
                            col_project = col_idx
                        if 'amount' in cell_val and col_amount is None:
                            col_amount = col_idx
                    
                    # Parse data rows (limit for performance)
                    for row_idx in range(header_row + 1, min(ws.max_row + 1, header_row + 1000)):
                        project_name = ws.cell(row_idx, col_project).value if col_project else None
                        if not project_name or not str(project_name).strip() or len(str(project_name).strip()) < 5:
                            continue
                        
                        # Try to find amount in various columns
                        amount = 0.0
                        if col_amount:
                            amount = self.parse_amount(ws.cell(row_idx, col_amount).value)
                        
                        # If no amount found, try other columns
                        if amount == 0:
                            for col_idx in range(col_project or 1, min(ws.max_column + 1, (col_project or 1) + 10)):
                                val = ws.cell(row_idx, col_idx).value
                                if isinstance(val, (int, float)) and val > 1000:  # Likely an amount
                                    amount = self.parse_amount(val)
                                    break
                        
                        if amount > 0:
                            self.proj_counter += 1
                            
                            projects_data.append({
                                "id": f"DPWH-Proj-{self.proj_counter:04d}",
                                "program_id": None,
                                "department_id": "DPWH",
                                "name": str(project_name).strip(),
                                "description": None,
                                "location": {"region": None, "province": None, "municipality": None, "barangay": None},
                                "original_amount": amount,
                                "final_amount": amount,
                                "net_change": 0.0,
                                "percent_change": 0.0,
                                "amendments": [],
                                "revised_name": None,
                                "page_reference": None,
                                "source_sheet": "DPWH 1-C ComRep"
                            })
            
            wb.close()
            print(f"✅ Parsed {len(projects_data)} DPWH projects")
            return projects_data
            
        except Exception as e:
            print(f"❌ Error parsing DPWH Projects: {e}")
            import traceback
            traceback.print_exc()
            return []

    
    def generate_json(self, output_path: str = "static/data/budget_amendments_2026.json"):
        """Generate final JSON output"""
        
        # Parse all files
        print("\n" + "="*60)
        print("🚀 Starting Budget Amendment Parsing")
        print("="*60 + "\n")
        
        # 1. Parse General Summary (departments)
        dept_data = self.parse_general_summary()
        self.departments = dept_data
        
        # 2. Parse Annex A (programs from 53 sheets)
        print()
        annex_a_programs = self.parse_annex_a()
        self.programs.extend(annex_a_programs)
        
        # 3. Parse DA Farm Roads (projects)
        print()
        farm_road_projects = self.parse_da_farm_roads()
        self.projects.extend(farm_road_projects)
        
        # 4. Parse Annex A-2 DepEd Schools (projects/programs)
        print()
        deped_schools = self.parse_deped_schools()
        self.projects.extend(deped_schools)
        
        # 5. Parse Annex A-4 NIA Operations (programs/projects)
        print()
        nia_operations = self.parse_nia_operations()
        self.programs.extend(nia_operations.get('programs', []))
        self.projects.extend(nia_operations.get('projects', []))
        
        # 6. Parse Annex A-5 DPWH Projects (projects)
        print()
        dpwh_projects = self.parse_dpwh_projects()
        self.projects.extend(dpwh_projects)
        
        # Detect aggregates and find their components
        self._detect_aggregates()
        
        # Validate and correct program-department assignments using NEP 2026 and GAB 2026
        print()
        print("🔍 Validating program-department assignments using NEP 2026 and GAB 2026...")
        self._validate_with_nep_gab()
        
        # Separate departments from agencies
        departments_list = []
        agencies_list = []
        
        for dept in self.departments:
            if dept.get('is_agency', False):
                agencies_list.append(dept)
            else:
                departments_list.append(dept)
        
        # Group agencies under their parent departments
        # Build a map of parent department names to department IDs
        parent_dept_map = {}
        for dept in departments_list:
            dept_name_upper = dept.get('name', '').upper()
            # Try to match parent names
            if 'DEPARTMENT OF' in dept_name_upper:
                parent_dept_map[dept_name_upper] = dept['id']
                # Also map without "DEPARTMENT OF" prefix
                short_name = dept_name_upper.replace('DEPARTMENT OF', '').strip()
                parent_dept_map[short_name] = dept['id']
        
        # Assign agencies to parent departments
        for agency in agencies_list:
            parent_name = agency.get('parent_department_name', '')
            if parent_name:
                parent_upper = parent_name.upper()
                # Try to find parent department
                parent_id = None
                if parent_upper in parent_dept_map:
                    parent_id = parent_dept_map[parent_upper]
                else:
                    # Try fuzzy match
                    for dept_name, dept_id in parent_dept_map.items():
                        if parent_upper in dept_name or dept_name in parent_upper:
                            parent_id = dept_id
                            break
                
                if parent_id:
                    agency['parent_department_id'] = parent_id
                    # Find parent department and add agency to it
                    parent_dept = next((d for d in departments_list if d['id'] == parent_id), None)
                    if parent_dept:
                        if 'agencies' not in parent_dept:
                            parent_dept['agencies'] = []
                        parent_dept['agencies'].append(agency)
        
        # Update department and agency program/project counts and amendment counts
        for dept in departments_list:
            dept['program_count'] = len([p for p in self.programs if p.get('department_id') == dept['id']])
            dept['project_count'] = len([p for p in self.projects if p.get('department_id') == dept['id']])
            # Count only meaningful amendments (items with actual changes)
            dept_line_items = [li for li in self.line_items 
                             if li.get('department_id') == dept['id']]
            meaningful_amendments = [
                li for li in dept_line_items
                if li.get('amendment_type') not in ['RETAIN', None]
            ]
            dept['amendment_count'] = len(meaningful_amendments)
            # Count agencies
            dept['agency_count'] = len(dept.get('agencies', []))
            
            # If department has 0 amounts but has line items, calculate amounts from line items
            # This handles cases where department was created from Annex A but amounts weren't set
            # BUT: Only do this if the department doesn't already have amounts from General Summary
            # General Summary amounts are authoritative for the full budget
            dept_original = dept.get('original_amount', 0)
            dept_final = dept.get('final_amount', 0)
            if dept_original == 0 and dept_final == 0 and dept_line_items:
                # Calculate from line items only if amounts are truly 0
                dept['original_amount'] = sum(li.get('original_amount', 0) for li in dept_line_items)
                dept['final_amount'] = sum(li.get('final_amount', 0) for li in dept_line_items)
                dept['increase'] = sum(li.get('increase', 0) for li in dept_line_items)
                dept['decrease'] = sum(li.get('decrease', 0) for li in dept_line_items)
                dept['net_change'] = dept['final_amount'] - dept['original_amount']
                dept['percent_change'] = (dept['net_change'] / dept['original_amount'] * 100) if dept['original_amount'] > 0 else 0.0
            elif dept_original > 0 or dept_final > 0:
                # Department has amounts from General Summary - keep them
                # Recalculate net_change and percent_change to ensure consistency
                dept['net_change'] = dept_final - dept_original
                dept['percent_change'] = (dept['net_change'] / dept_original * 100) if dept_original > 0 else 0.0
        
        for agency in agencies_list:
            agency['program_count'] = len([p for p in self.programs if p.get('department_id') == agency['id']])
            agency['project_count'] = len([p for p in self.projects if p.get('department_id') == agency['id']])
            agency_line_items = [li for li in self.line_items 
                                if li.get('department_id') == agency['id']]
            meaningful_amendments = [
                li for li in agency_line_items
                if li.get('amendment_type') not in ['RETAIN', None]
            ]
            agency['amendment_count'] = len(meaningful_amendments)
        
        # Calculate metadata
        metadata = {
            "fiscal_year": "2026",
            "bill_number": "HBN 4058",
            "generated_at": datetime.now().isoformat(),
            "total_departments": len(departments_list),
            "total_agencies": len(agencies_list),
            "total_programs": len(self.programs),
            "total_projects": len(self.projects),
            "total_original_budget": sum(d['original_amount'] for d in departments_list),
            "total_final_budget": sum(d['final_amount'] for d in self.departments),
            "total_net_change": sum(d['net_change'] for d in self.departments),
            "total_increase": sum(d['increase'] for d in self.departments),
            "total_decrease": sum(d['decrease'] for d in self.departments),
            "source_files": [
                "General Summary of Committee Report No. 18 on House Bill No. 4058 (FY 2026 GAB).xlsx",
                "Annex A - Line By Line Amendments.xlsx",
                "Annex A-1 DA-Farm-to-Market Roads.xlsx",
                "Annex A-2 DepEd - Office of the Secretary - Non Implementing Unit Secondary Schools.xlsx",
                "Annex A-4 BSGC-OEOs-NIA Details of NIA's Operations Budget.xlsx",
                "Annex A-5 Details of DPWH's Programs&Projects.xlsx"
            ]
        }
        
        # Group worksheets by department
        # Exclude SUCs worksheets from main view (they're huge components)
        # Group SUCs worksheets together
        worksheets_by_dept = {}
        sucs_worksheets = []
        
        for worksheet in self.worksheets:
            dept_id = worksheet.get('department_id', '')
            if worksheet.get('is_sucs', False):
                sucs_worksheets.append(worksheet)
            else:
                if dept_id not in worksheets_by_dept:
                    worksheets_by_dept[dept_id] = []
                worksheets_by_dept[dept_id].append(worksheet)
        
        # Add SUCs as a grouped component under DepEd
        if sucs_worksheets:
            deped_id = next((d['id'] for d in departments_list if 'EDUCATION' in d.get('name', '').upper() and 'OFFICE' not in d.get('name', '').upper()), None)
            if deped_id:
                # Create a combined SUCs entry
                sucs_entry = {
                    "id": "SUCS-GROUP",
                    "sheet_name": "SUCs",
                    "description": "State Universities and Colleges",
                    "department_id": deped_id,
                    "is_sucs": True,
                    "is_group": True,
                    "worksheet_count": len(sucs_worksheets),
                    "line_item_count": sum(ws.get('line_item_count', 0) for ws in sucs_worksheets),
                    "amendment_count": sum(ws.get('amendment_count', 0) for ws in sucs_worksheets),
                    "original_amount": sum(ws.get('original_amount', 0) for ws in sucs_worksheets),
                    "final_amount": sum(ws.get('final_amount', 0) for ws in sucs_worksheets),
                    "net_change": sum(ws.get('net_change', 0) for ws in sucs_worksheets),
                    "worksheets": sucs_worksheets  # Individual SUC worksheets
                }
                if deped_id not in worksheets_by_dept:
                    worksheets_by_dept[deped_id] = []
                worksheets_by_dept[deped_id].append(sucs_entry)
        
        # Add worksheets to departments
        for dept in departments_list:
            dept_id = dept.get('id', '')
            dept['worksheets'] = worksheets_by_dept.get(dept_id, [])
            dept['worksheet_count'] = len(dept['worksheets'])
        
        # Build final structure
        output = {
            "metadata": metadata,
            "departments": departments_list,  # Top-level departments with worksheets
            "agencies": agencies_list,  # Agencies (grouped under departments)
            "worksheets": self.worksheets,  # All worksheets (for reference)
            "programs": self.programs,
            "projects": self.projects,
            "line_items": self.line_items  # Detailed line-by-line amendments
        }
        
        # Update metadata
        metadata["total_line_items"] = len(self.line_items)
        
        # Write to file
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        
        # Print summary
        print("\n" + "="*60)
        print("✅ Budget Amendment Parsing Complete!")
        print("="*60)
        print(f"📁 Output: {output_path}")
        print(f"📊 Departments: {metadata['total_departments']}")
        print(f"📝 Programs: {metadata['total_programs']}")
        print(f"🎯 Projects: {metadata['total_projects']}")
        print(f"📋 Line Items: {metadata.get('total_line_items', 0)}")
        print(f"💰 Total Budget: ₱{metadata['total_final_budget'] / 1e12:.2f}T")
        print(f"📈 Net Change: ₱{metadata['total_net_change'] / 1e9:.2f}B")
        print(f"📄 Source Files: {len(metadata['source_files'])}")
        print("="*60 + "\n")
    
    def _detect_aggregates(self):
        """Detect aggregate entries and find their component departments"""
        # Group departments by parent department
        dept_by_parent = {}
        for dept in self.departments:
            parent = dept.get('parent_department')
            if parent:
                if parent not in dept_by_parent:
                    dept_by_parent[parent] = []
                dept_by_parent[parent].append(dept)
        
        # Check each department to see if it's an aggregate
        for dept in self.departments:
            if dept.get('is_aggregate'):
                # Find components (other departments with same parent, appearing after this one)
                parent = dept.get('parent_department')
                dept_row = dept.get('excel_row', 0)
                
                if parent and parent in dept_by_parent:
                    # Find departments with same parent that appear after this one
                    components = [
                        d for d in dept_by_parent[parent]
                        if d.get('excel_row', 0) > dept_row and d.get('id') != dept.get('id')
                    ]
                    
                    # Check if the sum of components matches this entry (within 1% tolerance)
                    if components:
                        component_sum = sum(c.get('original_amount', 0) for c in components)
                        dept_amount = dept.get('original_amount', 0)
                        
                        if dept_amount > 0 and abs(component_sum - dept_amount) / dept_amount < 0.01:
                            dept['is_aggregate'] = True
                            dept['component_ids'] = [c.get('id') for c in components]
                            dept['component_count'] = len(components)
                        else:
                            # Still mark as aggregate if keyword detected, but components might be elsewhere
                            dept['component_ids'] = []
                            dept['component_count'] = 0
                    else:
                        dept['component_ids'] = []
                        dept['component_count'] = 0
                else:
                    dept['component_ids'] = []
                    dept['component_count'] = 0
    
    def _similarity(self, str1: str, str2: str) -> float:
        """Calculate similarity between two strings (0-1)"""
        if not str1 or not str2:
            return 0.0
        return SequenceMatcher(None, str1.lower().strip(), str2.lower().strip()).ratio()
    
    def _normalize_name(self, name: str) -> str:
        """Normalize department/program name for matching"""
        if not name:
            return ""
        # Remove common prefixes/suffixes
        name = name.upper().strip()
        name = re.sub(r'^(A\.|B\.|C\.|D\.|E\.|F\.|G\.|H\.|I\.|J\.|K\.|L\.|M\.|N\.|O\.|P\.|Q\.|R\.|S\.|T\.|U\.|V\.|W\.|X\.|Y\.|Z\.)\s*', '', name)
        name = re.sub(r'\s+', ' ', name)
        return name.strip()
    
    def _validate_with_nep_gab(self):
        """Validate and correct program-department assignments using NEP 2026 and GAB 2026"""
        try:
            # Run async validation
            asyncio.run(self._async_validate_with_nep_gab())
        except Exception as e:
            print(f"⚠️  Error during validation: {e}")
            import traceback
            traceback.print_exc()
    
    async def _async_validate_with_nep_gab(self):
        """Async validation using NEP 2026 and GAB 2026 databases"""
        try:
            # Connect to NEP database
            nep_conn = await asyncpg.connect(
                host='localhost',
                port=5432,
                user='budget_admin',
                password='wuQ5gBYCKkZiOGb61chLcByMu',
                database='nep'
            )
            
            corrections = 0
            
            # 1. Get NEP 2026 program-department mappings
            # Use uacs_dpt_dsc from budget_analysis database for department names
            print("   📊 Loading NEP 2026 program-department mappings...")
            
            # First, try to get department descriptions from budget_analysis database
            try:
                budget_conn = await asyncpg.connect(
                    host='localhost',
                    port=5432,
                    user='budget_admin',
                    password='wuQ5gBYCKkZiOGb61chLcByMu',
                    database='budget_analysis'
                )
                
                # Get UACS department descriptions
                uacs_dept_map = {}
                uacs_rows = await budget_conn.fetch("""
                    SELECT DISTINCT uacs_dpt_code, uacs_dpt_dsc
                    FROM uacs_department
                    WHERE uacs_dpt_dsc IS NOT NULL
                """)
                for row in uacs_rows:
                    uacs_dept_map[row['uacs_dpt_code']] = row['uacs_dpt_dsc']
                
                await budget_conn.close()
            except Exception as e:
                print(f"      ⚠️  Could not load UACS department map: {e}")
                uacs_dept_map = {}
            
            # Get programs from NEP 2026
            nep_programs_raw = await nep_conn.fetch("""
                SELECT DISTINCT
                    description as program_name,
                    org_uacs_code,
                    COUNT(*) as item_count,
                    SUM(amount) as total_amount
                FROM budget_2026
                WHERE description IS NOT NULL 
                    AND description != ''
                    AND description NOT LIKE 'Project(s):%'
                    AND description NOT LIKE 'Locally-Funded Project(s):%'
                    AND description NOT LIKE 'Operations:%'
                    AND description NOT LIKE 'General Administration%'
                    AND description NOT LIKE 'Support to Operations%'
                    AND LENGTH(description) > 10
                GROUP BY description, org_uacs_code
                HAVING COUNT(*) >= 3  -- Only programs with multiple line items
                ORDER BY total_amount DESC
                LIMIT 500
            """)
            
            # Convert to list of dicts and add department name from UACS map
            nep_programs = []
            for row in nep_programs_raw:
                org_code = row['org_uacs_code'] or ''
                dept_name = ''
                # Extract department code (first 2 digits) and look up department name
                if org_code and len(org_code) >= 2:
                    dept_code = org_code[:2]
                    dept_name = uacs_dept_map.get(dept_code, '')
                
                nep_programs.append({
                    'program_name': row['program_name'],
                    'org_uacs_code': org_code,
                    'department': dept_name,
                    'item_count': row['item_count'],
                    'total_amount': row['total_amount']
                })
            
            print(f"   ✅ Loaded {len(nep_programs)} NEP 2026 programs")
            
            # 2. Get GAB 2026 department mappings
            print("   📊 Loading GAB 2026 department data...")
            gab_departments = await nep_conn.fetch("""
                SELECT DISTINCT
                    label as department_name,
                    original,
                    hgab,
                    delta
                FROM pbc_gab_2026_headings_detail
                WHERE label IS NOT NULL 
                    AND label != ''
                    AND label NOT LIKE 'TOTAL%'
                    AND label NOT LIKE 'DEPARTMENTS'
                ORDER BY original DESC
            """)
            
            print(f"   ✅ Loaded {len(gab_departments)} GAB 2026 departments")
            
            # 3. Build NEP program to department mapping
            nep_program_to_dept = {}
            for row in nep_programs:
                prog_name = self._normalize_name(row['program_name'])
                dept_name = row['department'] or ''
                org_code = row['org_uacs_code'] or ''
                
                if prog_name:
                    if prog_name not in nep_program_to_dept:
                        nep_program_to_dept[prog_name] = []
                    nep_program_to_dept[prog_name].append({
                        'department': dept_name,
                        'org_code': org_code,
                        'amount': float(row['total_amount'] or 0)
                    })
            
            # 4. Build GAB department name normalization map
            gab_dept_map = {}
            for row in gab_departments:
                dept_name = self._normalize_name(row['department_name'])
                if dept_name:
                    gab_dept_map[dept_name] = row['department_name']
            
            # 5. Build amendment department name map
            amendment_dept_map = {}
            for dept in self.departments:
                normalized = self._normalize_name(dept.get('name', ''))
                if normalized:
                    amendment_dept_map[normalized] = dept['id']
            
            # 6. Validate and correct program assignments
            print("   🔍 Validating program-department assignments...")
            
            checked = 0
            matches_found = 0
            
            for program in self.programs:
                prog_name = program.get('name', '')
                current_dept_id = program.get('department_id', '')
                normalized_prog = self._normalize_name(prog_name)
                
                if not normalized_prog or not prog_name:
                    continue
                
                checked += 1
                
                # Try to find matching program in NEP
                best_match = None
                best_score = 0.0
                best_dept = None
                
                for nep_prog_name, dept_list in nep_program_to_dept.items():
                    similarity = self._similarity(normalized_prog, nep_prog_name)
                    if similarity > best_score and similarity > 0.6:  # Lower threshold to 60%
                        best_score = similarity
                        best_match = nep_prog_name
                        # Use the department with highest amount
                        best_dept_info = max(dept_list, key=lambda x: x['amount'])
                        best_dept = best_dept_info['department']
                
                if best_match:
                    matches_found += 1
                
                # If we found a match, try to correct the department assignment
                if best_match and best_dept:
                    # Try to find matching department in amendments
                    normalized_gab_dept = self._normalize_name(best_dept)
                    
                    # Debug: show first few matches
                    if matches_found <= 3:
                        print(f"      🔍 Match #{matches_found}: '{prog_name[:50]}'")
                        print(f"         NEP program: '{best_match[:50]}' ({best_score:.2%})")
                        print(f"         NEP dept: '{best_dept[:50]}' (normalized: '{normalized_gab_dept[:50]}')")
                        print(f"         Current dept: {current_dept_id}")
                    
                    # Try exact match first
                    if normalized_gab_dept in amendment_dept_map:
                        correct_dept_id = amendment_dept_map[normalized_gab_dept]
                        if correct_dept_id != current_dept_id:
                            # Find the department name for logging
                            correct_dept_name = next(
                                (d['name'] for d in self.departments if d['id'] == correct_dept_id),
                                correct_dept_id
                            )
                            current_dept_name = next(
                                (d['name'] for d in self.departments if d['id'] == current_dept_id),
                                current_dept_id
                            )
                            
                            print(f"      ✅ Corrected: '{prog_name[:50]}'")
                            print(f"         From: {current_dept_name[:50]} ({current_dept_id})")
                            print(f"         To: {correct_dept_name[:50]} ({correct_dept_id})")
                            print(f"         Match: {best_score:.2%} with NEP program '{best_match[:50]}'")
                            
                            # Update program assignment
                            program['department_id'] = correct_dept_id
                            program['department_name'] = correct_dept_name
                            corrections += 1
                    else:
                        # Try fuzzy matching with amendment departments
                        best_dept_match = None
                        best_dept_score = 0.0
                        best_dept_id = None
                        
                        for norm_amd_dept, amd_dept_id in amendment_dept_map.items():
                            dept_similarity = self._similarity(normalized_gab_dept, norm_amd_dept)
                            if dept_similarity > best_dept_score and dept_similarity > 0.75:
                                best_dept_score = dept_similarity
                                best_dept_id = amd_dept_id
                        
                        if best_dept_id and best_dept_id != current_dept_id:
                            correct_dept_name = next(
                                (d['name'] for d in self.departments if d['id'] == best_dept_id),
                                best_dept_id
                            )
                            current_dept_name = next(
                                (d['name'] for d in self.departments if d['id'] == current_dept_id),
                                current_dept_id
                            )
                            
                            print(f"      ✅ Corrected (fuzzy): '{prog_name[:50]}'")
                            print(f"         From: {current_dept_name[:50]} ({current_dept_id})")
                            print(f"         To: {correct_dept_name[:50]} ({best_dept_id})")
                            print(f"         Program match: {best_score:.2%}, Dept match: {best_dept_score:.2%}")
                            
                            program['department_id'] = best_dept_id
                            program['department_name'] = correct_dept_name
                            corrections += 1
                
                # Also update line items for this program
                if program.get('id'):
                    for line_item in self.line_items:
                        if line_item.get('program_id') == program['id']:
                            line_item['department_id'] = program['department_id']
            
            await nep_conn.close()
            
            print(f"   📊 Checked {checked} programs, found {matches_found} NEP matches")
            print(f"   ✅ Validation complete: {corrections} program-department assignments corrected")
            
        except Exception as e:
            print(f"   ⚠️  Validation error: {e}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    parser = BudgetAmendmentParser()
    parser.generate_json()
